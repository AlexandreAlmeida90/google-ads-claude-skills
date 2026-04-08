#!/usr/bin/env python3
"""
B6 Title Optimizer — Part 2: AI-Powered Title Generation
Beyond Six — Google Ads Agency

Generates optimised product titles using the Claude API.

Workflow:
  Step 1 — Generate 15 sample titles for review and approval:
    python b6_title_optimizer.py --feed=products.tsv --brand=VitalPlanet --vertical=supplements --mode=sample --output=data/

  Step 2 — After approval, generate titles for flagged products or the full feed:
    python b6_title_optimizer.py --feed=products.tsv --brand=VitalPlanet --vertical=supplements --mode=full --scope=flagged --output=data/
    python b6_title_optimizer.py --feed=products.tsv --brand=VitalPlanet --vertical=supplements --mode=full --scope=all --output=data/

Requirements:
  pip install pandas openpyxl anthropic
  export ANTHROPIC_API_KEY=sk-ant-...
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Vertical definitions ───────────────────────────────────────────────────────

VERTICAL_KEYWORDS = {
    'apparel':     ['shirt', 'dress', 'pants', 'jacket', 'shoe', 'clothing', 'apparel',
                    'fashion', 'hoodie', 'socks', 'hat', 'jeans'],
    'supplements': ['probiotic', 'vitamin', 'supplement', 'capsule', 'softgel', 'enzyme',
                    'protein', 'collagen', 'omega', 'prebiotic'],
    'pet':         ['dog', 'cat', 'pet', 'puppy', 'kitten', 'bird', 'reptile',
                    'aquarium', 'paw', 'feline', 'canine'],
    'electronics': ['phone', 'laptop', 'camera', 'battery', 'cable', 'charger',
                    'speaker', 'headphone', 'tablet', 'monitor'],
    'home':        ['furniture', 'sofa', 'lamp', 'rug', 'shelf', 'decor',
                    'mattress', 'pillow', 'curtain', 'chair'],
    'food':        ['food', 'snack', 'beverage', 'drink', 'coffee', 'tea',
                    'organic', 'protein bar', 'granola'],
    'media':       ['book', 'dvd', 'game', 'software', 'music', 'album', 'film', 'blu-ray'],
    'automotive':  ['car', 'vehicle', 'auto', 'truck', 'motorcycle', 'tire', 'brake', 'motor'],
}

VERTICAL_LABELS = {
    'apparel':     'Apparel & Accessories',
    'supplements': 'Supplements & Health',
    'pet':         'Pet Supplies',
    'electronics': 'Electronics',
    'home':        'Home & Garden',
    'food':        'Food & Grocery',
    'media':       'Media & Entertainment',
    'automotive':  'Automotive',
    'general':     'General',
}

# Title pattern per vertical — drives the Claude system prompt
TITLE_PATTERNS = {
    'apparel':     'Brand + Gender + Product Type + Style/Material + Color + Size',
    'supplements': 'Brand + Product Line + Product Type + Count/Size + Key Benefit',
    'pet':         'Brand + Pet Type + Product Type + Key Feature + Count/Size',
    'electronics': 'Brand + Model + Product Type + Key Spec + Capacity',
    'home':        'Brand + Product Type + Material + Dimensions + Color',
    'food':        'Brand + Product Type + Variant/Flavour + Size/Weight/Count',
    'media':       'Title + Author/Artist + Format + Edition (if applicable)',
    'automotive':  'Brand + Vehicle Compatibility + Product Type + Key Spec',
    'general':     'Brand + Product Type + Key Attribute + Size/Color/Variant',
}

TITLE_RULES = """\
Rules for optimised Google Shopping titles:
- Front-load the most important keywords — Google reads left to right and truncates on the right
- Use " - " as the separator between major title components
- Target 70–150 characters (never exceed 150, aim for at least 50)
- Include the brand name if it is a recognisable search term
- Include specific searchable attributes: count, size, weight, flavour, colour, material, model number
- Banned words and phrases: Sale!, Best!, Free!, #1, Award-Winning, Limited Time, Exclusive, Guaranteed
- No ALL CAPS words
- No special characters except hyphens, commas, and ampersands
- Every title must be unique — include variant-distinguishing details (size, flavour, colour, count)
- Do not invent information not present in the product data provided"""

# ── Styling (matching b6_feed_audit.py) ───────────────────────────────────────

C_HEADER     = 'FF141D2E'
C_ACCENT     = 'FF3664F4'
C_WHITE      = 'FFFFFFFF'
C_LIGHT      = 'FFF8F9FF'
C_BORDER     = 'FFE5E7EB'
C_AMBER_FILL = 'FFFFF2CC'
C_GREEN_FILL = 'FFE8F5E9'
C_GREY_FILL  = 'FFF0F0F0'

_thin   = Side(style='thin', color=C_BORDER)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _hfont(bold=True, size=11, color=C_WHITE):
    return Font(name='Calibri', bold=bold, size=size, color=color)


def _center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)


def _left(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, row_num, headers, bg=C_ACCENT):
    for c, h in enumerate(headers, 1):
        cell            = ws.cell(row=row_num, column=c, value=h)
        cell.font       = _hfont(size=10)
        cell.fill       = _fill(bg)
        cell.border     = _border
        cell.alignment  = _center()


# ── Vertical detection ─────────────────────────────────────────────────────────

def detect_vertical(df):
    sample_text = ' '.join(
        df.get('title', pd.Series()).dropna().astype(str).head(50).tolist()
    ).lower()
    for vertical, keywords in VERTICAL_KEYWORDS.items():
        if any(k in sample_text for k in keywords):
            return vertical
    return 'general'


# ── Title issue detection ──────────────────────────────────────────────────────

def flag_title_issues(df):
    """Return list of df row indices with title problems (too short, too long, duplicates)."""
    if 'title' not in df.columns:
        return []
    flagged = set()
    lengths = df['title'].str.len().fillna(0)
    flagged.update(df[lengths < 25].index.tolist())
    flagged.update(df[lengths > 150].index.tolist())
    flagged.update(df[df['title'].duplicated(keep=False)].index.tolist())
    return sorted(flagged)


def _issue_label(row, flagged_set):
    if row.name not in flagged_set:
        return 'OK — can improve'
    title  = str(row.get('title', '') or '')
    length = len(title)
    if length < 25:
        return f'Too short ({length} chars)'
    if length > 150:
        return f'Too long ({length} chars)'
    return 'Duplicate title'


# ── Product payload for Claude ─────────────────────────────────────────────────

def _build_payload(row):
    """Extract the attributes Claude needs to write a great title."""
    attrs = {}
    for col in ['color', 'size', 'material', 'pattern', 'age_group', 'gender',
                'item_group_id', 'mpn', 'gtin']:
        val = row.get(col, None)
        if val is not None and str(val).strip() not in ('', 'nan'):
            attrs[col] = str(val).strip()

    desc = str(row.get('description', '') or '')
    if len(desc) > 300:
        desc = desc[:300] + '...'

    return {
        'id':                  str(row.get('id', '')),
        'current_title':       str(row.get('title', '') or ''),
        'description_excerpt': desc,
        'brand':               str(row.get('brand', '') or ''),
        'product_type':        str(row.get('product_type', '') or ''),
        'attributes':          attrs,
    }


# ── Claude API ─────────────────────────────────────────────────────────────────

def _system_prompt(vertical, brand):
    pattern = TITLE_PATTERNS.get(vertical, TITLE_PATTERNS['general'])
    label   = VERTICAL_LABELS.get(vertical, 'General')
    return (
        f'You are a Google Shopping feed specialist optimising titles for '
        f'{label} products by {brand}.\n\n'
        f'{TITLE_RULES}\n\n'
        f'Title structure for {label}:\n{pattern}\n\n'
        'You will receive a JSON array of products. Return a JSON array with one '
        'object per product:\n'
        '  "id"    — the product id (unchanged)\n'
        '  "title" — the new optimised title\n'
        '  "notes" — one short sentence describing the key change made\n\n'
        'Return ONLY valid JSON. No markdown, no explanation outside the JSON array.'
    )


def _call_claude(client, system_prompt, batch, model):
    """Call Claude with a batch of products. Returns a list of {id, title, notes} dicts."""
    response = client.messages.create(
        model=model,
        max_tokens=2048,
        system=system_prompt,
        messages=[{'role': 'user', 'content': json.dumps(batch, ensure_ascii=False)}],
    )
    content = response.content[0].text.strip()

    # Strip markdown code fences if present
    if '```' in content:
        for part in content.split('```'):
            part = part.strip().lstrip('json').strip()
            try:
                return json.loads(part)
            except json.JSONDecodeError:
                continue

    return json.loads(content)


def generate_titles(df_scope, client, system_prompt, model, batch_size=20):
    """
    Generate titles for all rows in df_scope.
    Returns dict: {product_id: {'title': ..., 'notes': ...}}
    """
    results = {}
    rows    = list(df_scope.iterrows())
    total   = len(rows)

    for start in range(0, total, batch_size):
        batch_rows = rows[start:start + batch_size]
        batch      = [_build_payload(row) for _, row in batch_rows]
        end        = min(start + batch_size, total)
        print(f'  Generating titles {start + 1}–{end} of {total}...')

        try:
            batch_results = _call_claude(client, system_prompt, batch, model)
            for item in batch_results:
                pid = str(item.get('id', ''))
                results[pid] = {
                    'title': str(item.get('title', '')).strip(),
                    'notes': str(item.get('notes', '')).strip(),
                }
        except Exception as e:
            print(f'  Warning: batch {start + 1}–{end} failed ({e}). Skipping.')

    return results


# ── Token / cost estimate ──────────────────────────────────────────────────────

def _cost_estimate(n_products, model):
    tokens_in  = n_products * 150   # ~150 input tokens per product
    tokens_out = n_products * 50    # ~50 output tokens per product
    if 'haiku' in model:
        cost = (tokens_in * 0.80 + tokens_out * 4.00) / 1_000_000
    elif 'sonnet' in model:
        cost = (tokens_in * 3.00 + tokens_out * 15.00) / 1_000_000
    else:
        cost = (tokens_in * 15.00 + tokens_out * 75.00) / 1_000_000
    return f'{tokens_in + tokens_out:,} tokens  (~${cost:.3f})'


# ── Excel: sample output ───────────────────────────────────────────────────────

def write_sample_excel(df_sample, results, flagged_set, out_path, vertical, brand):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Title Samples'
    ws.sheet_view.showGridLines = False

    # Header block
    ws.merge_cells('A1:G1')
    ws['A1'].value     = 'TITLE SAMPLES — REVIEW & APPROVE'
    ws['A1'].font      = _hfont(size=13)
    ws['A1'].fill      = _fill(C_HEADER)
    ws['A1'].alignment = _center()

    ws.merge_cells('A2:G2')
    ws['A2'].value     = (f'{brand}  ·  {VERTICAL_LABELS.get(vertical, vertical)}'
                          f'  ·  {datetime.now().strftime("%d %b %Y")}')
    ws['A2'].font      = Font(name='Calibri', size=10, color='FFFFFFFF')
    ws['A2'].fill      = _fill(C_HEADER)
    ws['A2'].alignment = _center()

    ws.merge_cells('A3:G3')
    ws['A3'].value     = ('Review the suggested titles below. Provide feedback if any adjustments '
                          'are needed before running the full generation.')
    ws['A3'].font      = Font(name='Calibri', size=10, color='FF2B303B', italic=True)
    ws['A3'].fill      = _fill('FFFFF8E1')
    ws['A3'].alignment = _left(wrap=True)
    ws.row_dimensions[3].height = 28

    _write_header_row(ws, 4, ['#', 'Product ID', 'Current Title', 'Chars', 'Issue', 'Suggested Title', 'New Chars'])

    for r, (_, row) in enumerate(df_sample.iterrows(), 5):
        pid        = str(row.get('id', ''))
        curr       = str(row.get('title', '') or '')
        result     = results.get(pid, {})
        suggested  = result.get('title', '— generation failed —')
        issue      = _issue_label(row, flagged_set)
        is_flagged = row.name in flagged_set

        ws.row_dimensions[r].height = 40

        cell_data = [
            (1, str(r - 4),     _center(),  _hfont(bold=True, size=10, color='FF2B303B'), C_LIGHT if r % 2 == 0 else C_WHITE),
            (2, pid,            _left(),    Font(name='Calibri', size=9,  color='FF737B8C'), C_LIGHT if r % 2 == 0 else C_WHITE),
            (3, curr,           _left(True), Font(name='Calibri', size=10, color='FF2B303B'), C_AMBER_FILL if is_flagged else (C_LIGHT if r % 2 == 0 else C_WHITE)),
            (4, len(curr),      _center(),  Font(name='Calibri', size=10, color='FF2B303B'), C_AMBER_FILL if is_flagged else (C_LIGHT if r % 2 == 0 else C_WHITE)),
            (5, issue,          _center(),  Font(name='Calibri', size=9,  color='FFE65100' if is_flagged else 'FF43A047'), C_LIGHT if r % 2 == 0 else C_WHITE),
            (6, suggested,      _left(True), Font(name='Calibri', size=10, color='FF1B5E20'), C_GREEN_FILL),
            (7, len(suggested), _center(),  Font(name='Calibri', size=10, color='FF2B303B'), C_GREEN_FILL),
        ]

        for col, val, align, font, bg in cell_data:
            cell           = ws.cell(row=r, column=col, value=val)
            cell.alignment = align
            cell.font      = font
            cell.fill      = _fill(bg)
            cell.border    = _border

    _set_col_widths(ws, [4, 18, 55, 7, 22, 55, 9])
    ws.freeze_panes = 'A5'
    wb.save(out_path)


# ── Excel: full output ─────────────────────────────────────────────────────────

def write_full_excel(df_all, df_scope, results, out_path, vertical, brand):
    wb    = Workbook()
    flagged_set = set(flag_title_issues(df_all))

    # ── Tab 1: Title Suggestions ───────────────────────────────────────────────
    ws1       = wb.active
    ws1.title = 'Title Suggestions'
    ws1.sheet_view.showGridLines = False

    n_generated = len(results)

    ws1.merge_cells('A1:G1')
    ws1['A1'].value     = 'TITLE OPTIMISATION REPORT'
    ws1['A1'].font      = _hfont(size=13)
    ws1['A1'].fill      = _fill(C_HEADER)
    ws1['A1'].alignment = _center()

    ws1.merge_cells('A2:G2')
    ws1['A2'].value     = (f'{brand}  ·  {VERTICAL_LABELS.get(vertical, vertical)}'
                           f'  ·  {datetime.now().strftime("%d %b %Y")}'
                           f'  ·  {n_generated} titles generated')
    ws1['A2'].font      = Font(name='Calibri', size=10, color='FFFFFFFF')
    ws1['A2'].fill      = _fill(C_HEADER)
    ws1['A2'].alignment = _center()

    _write_header_row(ws1, 3, ['#', 'Product ID', 'Current Title', 'Chars', 'Suggested Title', 'New Chars', 'Notes'])

    r = 4
    for _, row in df_scope.iterrows():
        pid        = str(row.get('id', ''))
        curr       = str(row.get('title', '') or '')
        result     = results.get(pid, {})
        suggested  = result.get('title', '— generation failed —')
        notes      = result.get('notes', '')
        is_flagged = row.name in flagged_set
        bg_row     = C_LIGHT if r % 2 == 0 else C_WHITE

        ws1.row_dimensions[r].height = 35

        cell_data = [
            (1, str(r - 3),     _center(),   _hfont(bold=True, size=10, color='FF2B303B'), bg_row),
            (2, pid,            _left(),     Font(name='Calibri', size=9, color='FF737B8C'), bg_row),
            (3, curr,           _left(True), Font(name='Calibri', size=10, color='FF2B303B'), C_AMBER_FILL if is_flagged else bg_row),
            (4, len(curr),      _center(),   Font(name='Calibri', size=10, color='FF2B303B'), C_AMBER_FILL if is_flagged else bg_row),
            (5, suggested,      _left(True), Font(name='Calibri', size=10, color='FF1B5E20'), C_GREEN_FILL),
            (6, len(suggested), _center(),   Font(name='Calibri', size=10, color='FF2B303B'), C_GREEN_FILL),
            (7, notes,          _left(True), Font(name='Calibri', size=9, color='FF737B8C'), bg_row),
        ]

        for col, val, align, font, bg in cell_data:
            cell           = ws1.cell(row=r, column=col, value=val)
            cell.alignment = align
            cell.font      = font
            cell.fill      = _fill(bg)
            cell.border    = _border

        r += 1

    _set_col_widths(ws1, [4, 18, 52, 7, 52, 9, 42])
    ws1.freeze_panes = 'A4'

    # ── Tab 2: Updated Feed ────────────────────────────────────────────────────
    ws2       = wb.create_sheet('Updated Feed')
    ws2.sheet_view.showGridLines = False

    cols      = list(df_all.columns)
    title_col = cols.index('title') + 1 if 'title' in cols else None

    _write_header_row(ws2, 1, cols, bg=C_HEADER)

    for r2, (_, row) in enumerate(df_all.iterrows(), 2):
        pid       = str(row.get('id', ''))
        new_title = results.get(pid, {}).get('title')

        for c, col in enumerate(cols, 1):
            raw = row.get(col, '')
            val = '' if (raw is None or str(raw) == 'nan') else str(raw)

            if c == title_col and new_title:
                val       = new_title
                cell_fill = _fill(C_GREEN_FILL)
            else:
                cell_fill = _fill(C_LIGHT if r2 % 2 == 0 else C_WHITE)

            cell           = ws2.cell(row=r2, column=c, value=val)
            cell.font      = Font(name='Calibri', size=9, color='FF2B303B')
            cell.fill      = cell_fill
            cell.border    = _border
            cell.alignment = _left()

    for c, col in enumerate(cols, 1):
        if col == 'id':
            ws2.column_dimensions[get_column_letter(c)].width = 15
        elif col == 'title':
            ws2.column_dimensions[get_column_letter(c)].width = 60
        else:
            ws2.column_dimensions[get_column_letter(c)].width = 18

    ws2.freeze_panes = 'A2'
    wb.save(out_path)


# ── Supplemental TSV ───────────────────────────────────────────────────────────

def write_supplemental_tsv(df_all, results, out_path):
    """
    Write a 2-column supplemental feed: id + updated title.
    Upload this to Merchant Center as a supplemental feed to override titles
    without re-uploading the entire primary feed.
    """
    rows = []
    for _, row in df_all.iterrows():
        pid       = str(row.get('id', ''))
        new_title = results.get(pid, {}).get('title')
        if new_title:
            rows.append({'id': pid, 'title': new_title})
    if rows:
        pd.DataFrame(rows).to_csv(out_path, sep='\t', index=False)
    return len(rows)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='B6 Title Optimizer — Part 2')
    parser.add_argument('--feed',     required=True,
                        help='Path to the TSV feed file')
    parser.add_argument('--brand',    required=True,
                        help='Brand name (e.g. VitalPlanet)')
    parser.add_argument('--vertical', default=None,
                        help='Override vertical (supplements, apparel, electronics, home, food, pet, automotive, media, general)')
    parser.add_argument('--mode',     required=True,  choices=['sample', 'full'],
                        help='sample = generate 15 titles for approval; full = generate all')
    parser.add_argument('--scope',    default='flagged', choices=['flagged', 'all'],
                        help='full mode: flagged = only products with title issues; all = entire feed')
    parser.add_argument('--output',   default='.',
                        help='Output directory (default: current directory)')
    parser.add_argument('--model',    default='claude-haiku-4-5-20251001',
                        help='Claude model to use (default: claude-haiku-4-5-20251001)')
    parser.add_argument('--api-key',  default=None,
                        help='Anthropic API key (or set ANTHROPIC_API_KEY env var)')
    args = parser.parse_args()

    # ── Load feed ──────────────────────────────────────────────────────────────
    print('\nB6 TITLE OPTIMIZER — Part 2')
    print(f'Brand:    {args.brand}')
    print(f'Feed:     {args.feed}')

    df = pd.read_csv(args.feed, sep='\t', dtype=str).fillna('')
    print(f'Products: {len(df)}')

    vertical = args.vertical or detect_vertical(df)
    print(f'Vertical: {VERTICAL_LABELS.get(vertical, vertical)}')

    # ── Flag title issues ──────────────────────────────────────────────────────
    flagged_ids_list = flag_title_issues(df)
    flagged_set      = set(flagged_ids_list)
    print(f'Flagged:  {len(flagged_set)} products with title issues '
          f'(too short / too long / duplicate)')

    # ── Init Claude client ─────────────────────────────────────────────────────
    try:
        import anthropic
    except ImportError:
        print('\nError: anthropic package not installed.')
        print('Run:   pip install anthropic')
        sys.exit(1)

    api_key = args.api_key or os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        print('\nError: Anthropic API key not found.')
        print('Set the ANTHROPIC_API_KEY environment variable or pass --api-key.')
        sys.exit(1)

    client        = anthropic.Anthropic(api_key=api_key)
    system_prompt = _system_prompt(vertical, args.brand)
    out_dir       = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)
    date_str      = datetime.now().strftime('%Y%m%d')
    brand_slug    = args.brand.lower().replace(' ', '')

    # ── Sample mode ────────────────────────────────────────────────────────────
    if args.mode == 'sample':
        # Pick up to 15 products — prioritise flagged, fill from top of feed
        sample_idx = flagged_ids_list[:15]
        if len(sample_idx) < 15:
            unflagged   = [i for i in df.index if i not in flagged_set]
            sample_idx += unflagged[:15 - len(sample_idx)]
        sample_idx = sample_idx[:20]

        df_sample = df.loc[sample_idx]
        print(f'\nGenerating {len(df_sample)} sample titles (1 API call)...')

        results  = generate_titles(df_sample, client, system_prompt, args.model)
        out_path = out_dir / f'{date_str}-{brand_slug}-title-samples.xlsx'

        write_sample_excel(df_sample, results, flagged_set, out_path, vertical, args.brand)

        print(f'\nSample titles saved to: {out_path}')
        print()
        print('── Next steps ──────────────────────────────────────────────')
        print('1. Open the Excel file and review the suggested titles.')
        print('2. Provide feedback if adjustments are needed.')
        print('3. When happy, run with --mode=full and choose your scope:')
        print()
        print(f'   --scope=flagged  {len(flagged_set):>5} products   {_cost_estimate(len(flagged_set), args.model)}')
        print(f'   --scope=all      {len(df):>5} products   {_cost_estimate(len(df), args.model)}')
        print()
        print('Note: flagged = only products with title issues (too short / duplicate / too long).')
        print('      all     = entire feed — recommended if you want to improve every title.')

    # ── Full mode ──────────────────────────────────────────────────────────────
    elif args.mode == 'full':
        if args.scope == 'flagged':
            df_scope = df.loc[flagged_ids_list] if flagged_ids_list else df.iloc[0:0]
            scope_label = f'flagged products ({len(df_scope)})'
        else:
            df_scope    = df
            scope_label = f'full feed ({len(df_scope)} products)'

        print(f'\nScope:  {scope_label}')
        print(f'Model:  {args.model}')
        print(f'Est:    {_cost_estimate(len(df_scope), args.model)}')
        print()

        results = generate_titles(df_scope, client, system_prompt, args.model)

        # Excel report
        out_excel = out_dir / f'{date_str}-{brand_slug}-titles.xlsx'
        write_full_excel(df, df_scope, results, out_excel, vertical, args.brand)

        # Supplemental TSV for direct Merchant Center upload
        out_tsv = out_dir / f'{date_str}-{brand_slug}-title-feed.tsv'
        n_tsv   = write_supplemental_tsv(df, results, out_tsv)

        print()
        print(f'Title optimisation complete.')
        print(f'  {len(results)} titles generated')
        print(f'  Excel report:      {out_excel}')
        print(f'  Supplemental feed: {out_tsv}  ({n_tsv} rows)')
        print()
        print('The supplemental feed (id + title) can be uploaded directly to')
        print('Merchant Center as a supplemental feed to override titles without')
        print('re-uploading your entire primary feed.')


if __name__ == '__main__':
    main()
