#!/usr/bin/env python3
"""
B6 Product Feed Audit — Phase 1: Attribute Audit
Beyond Six — Google Ads Agency

Outputs a 3-tab Excel file:
  Tab 1: Issues Summary — one row per specific issue, ranked by priority
  Tab 2: Attribute Audit — all required + recommended attributes with status
  Tab 3: All Products — full feed with cell-level color coding

Usage:
  python b6_feed_audit.py --feed=products.tsv --brand=BrightSkin
  python b6_feed_audit.py --feed=products.tsv --brand=BrightSkin --vertical=supplements
  python b6_feed_audit.py --feed=products.tsv --brand=BrightSkin --output=data/feeds/
  python b6_feed_audit.py --feed=products.tsv --brand=BrightSkin --detect-only
"""

import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Vertical definitions ───────────────────────────────────────────────────────

VERTICAL_CATEGORY_IDS = {
    # Specific verticals first — checked before broader parent categories
    'watches':        {'201', '5122', '5123'},
    'jewelry':        {'188', '189', '190', '191', '192', '194', '196', '197', '200', '6463'},
    'footwear':       {'187', '5387'},
    'bags':           {'5181'},
    'beauty':         {'473', '2915'},
    'furniture':      {'436'},
    'sporting_goods': {'988'},
    'toys':           {'1239'},
    'baby':           {'537'},
    # Broader verticals
    'apparel':      {'1604', '166', '5697', '2271', '178'},
    'supplements':  {'525', '5413'},
    'pet':          {'1', '2', '3', '4', '5', '5081', '5086'},
    'electronics':  {'222', '412'},
    'home':         {'536', '638', '594'},
    'food':         {'422', '5765'},
    'media':        {'783', '839', '855'},
    'automotive':   {'913'},
}

VERTICAL_KEYWORDS = {
    # Specific verticals first — matched before broader parent categories
    'watches':        ['watch', 'chronograph', 'quartz', 'automatic', 'timepiece', 'wristwatch', 'smartwatch', 'horology'],
    'jewelry':        ['ring', 'necklace', 'bracelet', 'earring', 'pendant', 'diamond', 'gemstone', 'brooch', 'anklet', 'jewellery', 'jewelry'],
    'footwear':       ['shoe', 'boot', 'sneaker', 'sandal', 'heel', 'loafer', 'trainer', 'stiletto', 'moccasin', 'slipper', 'footwear'],
    'bags':           ['bag', 'handbag', 'backpack', 'wallet', 'purse', 'tote', 'clutch', 'luggage', 'suitcase', 'briefcase', 'duffel'],
    'beauty':         ['serum', 'foundation', 'lipstick', 'mascara', 'cleanser', 'toner', 'shampoo', 'conditioner', 'moisturiser', 'moisturizer', 'makeup', 'concealer', 'blush', 'eyeshadow'],
    'furniture':      ['sofa', 'couch', 'wardrobe', 'cabinet', 'dresser', 'ottoman', 'bookcase', 'nightstand', 'armchair', 'sideboard'],
    'sporting_goods': ['gym', 'fitness', 'yoga', 'cycling', 'tennis', 'football', 'golf', 'swimming', 'hiking', 'workout', 'barbell', 'dumbbell', 'racket', 'treadmill'],
    'toys':           ['toy', 'puzzle', 'lego', 'doll', 'board game', 'playset', 'action figure', 'stuffed animal', 'building blocks'],
    'baby':           ['baby', 'infant', 'toddler', 'newborn', 'diaper', 'stroller', 'crib', 'onesie', 'nursery', 'pacifier', 'pram'],
    # Broader verticals
    'apparel':     ['shirt', 'dress', 'pants', 'jacket', 'clothing', 'apparel', 'fashion', 'hoodie', 'socks', 'hat', 'jeans'],
    'supplements': ['probiotic', 'vitamin', 'supplement', 'capsule', 'softgel', 'enzyme', 'protein', 'collagen', 'omega', 'prebiotic'],
    'pet':         ['dog', 'cat', 'pet', 'puppy', 'kitten', 'bird', 'reptile', 'aquarium', 'paw', 'feline', 'canine'],
    'electronics': ['phone', 'laptop', 'camera', 'battery', 'cable', 'charger', 'speaker', 'headphone', 'tablet', 'monitor'],
    'home':        ['lamp', 'rug', 'shelf', 'decor', 'mattress', 'pillow', 'curtain', 'chair', 'furniture'],
    'food':        ['food', 'snack', 'beverage', 'drink', 'coffee', 'tea', 'organic', 'protein bar', 'granola'],
    'media':       ['book', 'dvd', 'game', 'software', 'music', 'album', 'film', 'blu-ray'],
    'automotive':  ['car', 'vehicle', 'auto', 'truck', 'motorcycle', 'tire', 'brake', 'motor'],
}

REQUIRED_UNIVERSAL = [
    'id', 'title', 'description', 'link', 'image_link',
    'availability', 'price', 'condition',
    'brand', 'gtin', 'identifier_exists', 'google_product_category',
]

REQUIRED_EXTRA = {
    # Watches: item_group_id for variant grouping (strap colour, dial colour)
    'watches':        ['item_group_id'],
    # Jewelry: material and colour are primary search attributes; item_group_id for size/metal variants
    'jewelry':        ['item_group_id', 'material', 'color'],
    # Footwear: same as apparel — size, colour, gender, age_group are required by Google
    'footwear':       ['item_group_id', 'size', 'color', 'gender', 'age_group'],
    # Bags: material is a primary search attribute; item_group_id for colour/size variants
    'bags':           ['item_group_id', 'material'],
    # Beauty: item_group_id for shade/scent/size variants
    'beauty':         ['item_group_id'],
    # Furniture: no extra required attrs — Google does not mandate demographic/variant attrs
    'furniture':      [],
    # Sporting goods: no extra required attrs — sport type captured in title/product_type
    'sporting_goods': [],
    # Toys: age_group required — Google uses it for filtering in toy search results
    'toys':           ['age_group'],
    # Baby: age_group required — primary filter for all baby/toddler products
    'baby':           ['age_group'],
    # Apparel: variants + demographic attributes are required by Google
    'apparel':     ['item_group_id', 'age_group', 'gender', 'color', 'size'],
    # Supplements: variants expected (count, size, flavor)
    'supplements': ['item_group_id'],
    'pet':         [],
    # Electronics: mpn required when GTIN unavailable — very common
    'electronics': ['mpn'],
    'home':        [],
    'food':        [],
    'media':       [],
    # Automotive: mpn is the primary product identifier for parts
    'automotive':  ['mpn'],
    'general':     [],
}

RECOMMENDED_CORE = [
    # Content & discoverability
    'product_type',
    'additional_image_link', 'lifestyle_image_link', 'mobile_link',
    'product_highlight', 'product_detail',
    # Pricing
    'sale_price', 'sale_price_effective_date',
    # Availability
    'availability_date',
    # Identifiers
    'mpn',
    # Shipping & returns
    'shipping_label', 'return_policy_label',
    # Performance & campaign management
    'cost_of_goods_sold',
    'custom_label_0', 'custom_label_1', 'custom_label_2', 'custom_label_3', 'custom_label_4',
    # Variants & bundles (conditional — only recommended when feed signals are found)
    'item_group_id', 'is_bundle', 'multipack',
    # Physical & demographic attributes (conditional — only recommended when feed signals are found)
    'gender', 'color', 'size', 'material', 'pattern', 'age_group',
]

RECOMMENDED_EXTRA = {
    # Watches: size_system for case diameter (mm); material for case/strap; colour for dial/strap
    'watches':        ['size_system', 'material', 'color'],
    # Jewelry: ring size, metal purity; pattern for engraved/textured designs
    'jewelry':        ['size', 'size_system', 'pattern'],
    # Footwear: size system (US/EU/UK), size type (wide/narrow), material (leather/mesh), pattern
    'footwear':       ['size_system', 'size_type', 'material', 'pattern'],
    # Bags: colour, size (dimensions/capacity), pattern (logo/plain)
    'bags':           ['color', 'size', 'pattern'],
    # Beauty: material is sometimes used for key active ingredients
    'beauty':         ['material'],
    # Furniture: material (wood/metal/fabric), colour, size (dimensions — W×D×H)
    'furniture':      ['material', 'color', 'size'],
    # Sporting goods: material, colour, size, plus demographic filters
    'sporting_goods': ['material', 'color', 'size', 'gender', 'age_group'],
    # Toys: material (safety relevance), colour, size, bundle/multipack signals
    'toys':           ['material', 'color', 'size', 'is_bundle', 'multipack'],
    # Baby: material (BPA-free/organic), size (clothing), gender, size system
    'baby':           ['material', 'size', 'gender', 'color', 'size_system'],
    # Apparel: size system/type are structural metadata for apparel search filtering
    # color/size/gender/age_group are REQUIRED for apparel (in REQUIRED_EXTRA above)
    # material/pattern promoted to conditional (feed-signal based) — no longer apparel-only
    'apparel':     ['size_system', 'size_type'],
    # Supplements: no vertical-specific extras — conditional attrs handle gender/size signals
    'supplements': [],
    # Pet: no vertical-specific extras — conditional attrs handle age/color/size/material signals
    'pet':         [],
    # Electronics: EU energy labels + certification are regulatory, not signal-based
    'electronics': ['energy_efficiency_class',
                    'min_energy_efficiency_class', 'max_energy_efficiency_class', 'certification'],
    # Home: energy labels for appliances; unit pricing required in EU food/household
    'home':        ['energy_efficiency_class',
                    'unit_pricing_measure', 'unit_pricing_base_measure'],
    # Food: unit pricing required in EU; expiration date for perishables
    'food':        ['unit_pricing_measure', 'unit_pricing_base_measure', 'expiration_date'],
    # Media: no physical attributes
    'media':       [],
    # Automotive: certification for regulated parts
    'automotive':  ['certification'],
    # General: no extras — conditional attrs cover all relevant physical signals
    'general':     [],
}

# ── Master attribute list — ALL known Google Shopping attributes ───────────────
# Used to make the Attribute Audit tab exhaustive: every attribute is listed,
# even ones not recommended for the current vertical (shown as "Not Recommended").

ALL_ATTRIBUTES = [
    # ── Basic product data ────────────────────────────────────────────────────
    'id', 'title', 'description', 'link', 'mobile_link',
    'image_link', 'additional_image_link', 'lifestyle_image_link',
    # ── Price ─────────────────────────────────────────────────────────────────
    'price', 'sale_price', 'sale_price_effective_date',
    'auto_pricing_min_price',
    'unit_pricing_measure', 'unit_pricing_base_measure',
    'installment', 'subscription_cost', 'loyalty_points',
    # ── Availability ──────────────────────────────────────────────────────────
    'availability', 'availability_date', 'expiration_date',
    # ── Product identifiers ───────────────────────────────────────────────────
    'brand', 'gtin', 'mpn', 'identifier_exists', 'certification',
    # ── Category ──────────────────────────────────────────────────────────────
    'google_product_category', 'product_type',
    # ── Product details ───────────────────────────────────────────────────────
    'condition', 'adult', 'multipack', 'is_bundle',
    'energy_efficiency_class', 'min_energy_efficiency_class', 'max_energy_efficiency_class',
    'age_group', 'color', 'gender', 'material', 'pattern',
    'size', 'size_system', 'size_type',
    'item_group_id', 'product_detail', 'product_highlight',
    # ── Campaign management ───────────────────────────────────────────────────
    'ads_grouping', 'ads_labels', 'ads_redirect',
    'custom_label_0', 'custom_label_1', 'custom_label_2', 'custom_label_3', 'custom_label_4',
    # ── Destinations ──────────────────────────────────────────────────────────
    'excluded_destination', 'included_destination',
    'shopping_ads_excluded_country', 'pause',
    # ── Shipping ──────────────────────────────────────────────────────────────
    'shipping', 'shipping_label', 'shipping_weight',
    'shipping_length', 'shipping_width', 'shipping_height',
    'ships_from_country', 'min_handling_time', 'max_handling_time',
    # ── Returns ───────────────────────────────────────────────────────────────
    'return_policy_label', 'return_policy_country',
    # ── Tax ───────────────────────────────────────────────────────────────────
    'tax', 'tax_category',
    # ── Cost ──────────────────────────────────────────────────────────────────
    'cost_of_goods_sold',
]

VERTICAL_LABELS = {
    'watches':        'Watches & Accessories',
    'jewelry':        'Jewelry & Accessories',
    'footwear':       'Footwear',
    'bags':           'Bags & Luggage',
    'beauty':         'Beauty & Personal Care',
    'furniture':      'Furniture & Home',
    'sporting_goods': 'Sporting Goods & Fitness',
    'toys':           'Toys & Games',
    'baby':           'Baby & Toddler',
    'apparel':        'Apparel & Accessories',
    'supplements':    'Supplements & Health',
    'pet':            'Pet Supplies',
    'electronics':    'Electronics',
    'home':           'Home & Garden',
    'food':           'Food & Grocery',
    'media':          'Media & Entertainment',
    'automotive':     'Automotive',
    'general':        'General',
}

# Valid values for key attributes
VALID_VALUES = {
    'identifier_exists':       {'yes', 'no', 'true', 'false'},
    'availability':            {'in stock', 'out of stock', 'preorder', 'backorder'},
    'condition':               {'new', 'refurbished', 'used'},
    'energy_efficiency_class': {'a+++', 'a++', 'a+', 'a', 'b', 'c', 'd', 'e', 'f', 'g'},
    'min_energy_efficiency_class': {'a+++', 'a++', 'a+', 'a', 'b', 'c', 'd', 'e', 'f', 'g'},
    'max_energy_efficiency_class': {'a+++', 'a++', 'a+', 'a', 'b', 'c', 'd', 'e', 'f', 'g'},
}

# What to do instructions per attribute
WHAT_TO_DO = {
    'id':                    'Every product must have a unique ID. Usually auto-generated by your platform.',
    'title':                 'Add a descriptive title: Brand + Product Type + Key Attribute (size, colour, count). Target 70–150 chars.',
    'description':           'Write a unique description listing key attributes and specs. No marketing copy. Target 500–2000 chars.',
    'link':                  'Add the product page URL. Must be a live, crawlable page.',
    'image_link':            'Add the main product image URL. White background preferred, no watermarks, min 800×800px.',
    'availability':          'Set to: in stock | out of stock | preorder | backorder.',
    'price':                 'Add price with currency code (e.g. 29.99 USD). Must match the price on the product page.',
    'condition':             'Set to: new | refurbished | used.',
    'brand':                 'Add the brand name. Required for most product types.',
    'gtin':                  'Add the GTIN (barcode). For products without one, set identifier_exists=no instead.',
    'identifier_exists':     'Set to yes if product has a GTIN or MPN. Set to no for custom/handmade products with no unique identifier.',
    'google_product_category': "Use the most specific subcategory from Google's taxonomy. More specific = better match quality and reach.",
    'item_group_id':         'Assign a shared item_group_id to all variants of the same base product (colour, size, etc.).',
    'age_group':             'Set to: newborn | infant | toddler | kids | adult.',
    'gender':                'Set to: male | female | unisex.',
    'color':                 'Add the primary colour(s) of the product.',
    'size':                  'Add the size using standardised values for the target market.',
    'mpn':                   'Add the Manufacturer Part Number. Required when GTIN is unavailable.',
    'sale_price':            'Add the sale price if the product is on promotion (e.g. 19.99 USD). Leave blank if not on sale.',
    'additional_image_link': 'Add extra product image URLs — different angles, close-ups. Up to 10 additional images allowed.',
    'lifestyle_image_link':  'Add a lifestyle/in-use image URL showing the product in context.',
    'product_type':          'Add your own product taxonomy using ">" as the separator, going from broad to specific (e.g. Supplements > Probiotics > Women\'s Probiotics > 50 Billion CFU Capsules). Google uses this attribute to semantically understand what you\'re selling — the more specific, the better.',
    'product_highlight':     'Add 2–5 short benefit-focused bullet points. These appear as bullets in Shopping listings.',
    'product_detail':        'Add technical specifications as attribute/value pairs (e.g. Serving Size: 2 capsules).',
    'shipping_label':        'Add a label to group products with the same shipping rules (e.g. standard, oversized, free-shipping).',
    'cost_of_goods_sold':    'Add the cost of goods sold per product. Enables gross profit reporting and margin-based bidding in Google Ads.',
    'is_bundle':             'Set to yes if the product is a bundle of multiple distinct items sold together.',
    'multipack':             'Set to the number of units if this is a multipack (e.g. 3 for a 3-pack).',
    'material':              'Add the primary material (e.g. cotton, stainless steel, silicone).',
    'pattern':               'Add the pattern if applicable (e.g. striped, floral, solid).',
    'size_system':           'Add the sizing system used (e.g. US, EU, UK, IT).',
    'size_type':             'Add the size type (e.g. regular, plus, petite, tall, maternity).',
    # Pricing
    'sale_price_effective_date': 'Add the date range for the sale price (e.g. 2024-01-01T00:00-08:00/2024-01-31T23:59-08:00). Without it, sale prices may not display correctly.',
    # Availability
    'availability_date':     'Add the date the product becomes available (required when availability=preorder). Format: ISO 8601 (e.g. 2024-06-01T00:00-08:00).',
    'expiration_date':       'Add the date the product listing should expire. Useful for seasonal or perishable products. Format: ISO 8601.',
    # Images & links
    'mobile_link':           'Add a mobile-optimised landing page URL if different from the desktop link. Improves mobile conversion tracking.',
    # Returns
    'return_policy_label':   'Add a return policy label matching a policy configured in Merchant Center (e.g. free-returns, no-returns). Google surfaces return info in listings.',
    # Identifiers
    'certification':         'Add the product certification (e.g. EU energy label: EC:A+). Required in the EU for energy-related products.',
    # Energy efficiency (EU regulated)
    'energy_efficiency_class':     "Add the EU energy efficiency rating (e.g. A+++, A++, A+, A, B, C, D, E, F, G). Required for regulated product categories in the EU.",
    'min_energy_efficiency_class': "Add the minimum energy efficiency class if the product spans a range (e.g. D). Used together with max_energy_efficiency_class.",
    'max_energy_efficiency_class': "Add the maximum energy efficiency class if the product spans a range (e.g. A+). Used together with min_energy_efficiency_class.",
    # Unit pricing (EU food/household)
    'unit_pricing_measure':      'Add the unit measure for price comparison (e.g. 1 kg, 750 ml). Required in several EU markets for food, beverage, and household products.',
    'unit_pricing_base_measure': 'Add the base measure for unit price display (e.g. 100g, 1L). Must be consistent with unit_pricing_measure.',
    # Campaign management
    'custom_label_0': 'Use for campaign segmentation — typically margin tier (e.g. high / medium / low). Requires data from your commerce platform.',
    'custom_label_1': 'Use for performance tier (e.g. bestseller / regular / new). Pull from Shopify or analytics top-sellers data.',
    'custom_label_2': 'Use for seasonal relevance (e.g. evergreen / spring / summer / fall / winter).',
    'custom_label_3': 'Use for price bracket (e.g. under-25 / 25-75 / 75-150 / 150-plus). Derived from the price column.',
    'custom_label_4': 'Use for promo eligibility (e.g. promo-eligible / excluded). Useful for controlling which products enter promotion campaigns.',
}


def _what_to_do(attr, context='incomplete'):
    return WHAT_TO_DO.get(attr, f'Add or correct the {attr} value on all affected products.')


# ── Colours ────────────────────────────────────────────────────────────────────

C_HEADER    = 'FF141D2E'
C_SUB       = 'FF2B303B'
C_ACCENT    = 'FF3664F4'
C_WHITE     = 'FFFFFFFF'
C_LIGHT     = 'FFF8F9FF'
C_BORDER    = 'FFE5E7EB'
C_GREY_FILL = 'FFF0F0F0'   # non-audited columns in All Products

# Cell fill colours (light enough to read text over)
C_RED_FILL   = 'FFFFCCCC'  # required attr missing/invalid
C_AMBER_FILL = 'FFFFF2CC'  # recommended attr missing, or invalid value
C_GREEN_TEXT = 'FF2E7D32'
C_AMBER_TEXT = 'FFE65100'
C_RED_TEXT   = 'FFC62828'

thin   = Side(style='thin', color=C_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def hfont(bold=True, size=11, color=C_WHITE):
    return Font(name='Calibri', bold=bold, size=size, color=color)


def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)


def left(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)


def set_col_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_header_row(ws, row_num, values, bg=C_HEADER, fg=C_WHITE, size=11):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = Font(name='Calibri', bold=True, size=size, color=fg)
        cell.fill = fill(bg)
        cell.alignment = center(wrap=True)
        cell.border = border


# ── Taxonomy ───────────────────────────────────────────────────────────────────

TAXONOMY_PATH = Path(__file__).parent.parent / 'google-product-taxonomy.txt'


def load_taxonomy():
    taxonomy = {}
    if TAXONOMY_PATH.exists():
        with open(TAXONOMY_PATH, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if ' - ' in line and not line.startswith('#'):
                    parts = line.split(' - ', 1)
                    if parts[0].isdigit():
                        taxonomy[parts[0]] = parts[1]
    return taxonomy


TAXONOMY = load_taxonomy()


def decode_category(val):
    if not val or pd.isna(val):
        return ''
    val = str(val).strip()
    return TAXONOMY.get(val, val)


# ── Load feed ──────────────────────────────────────────────────────────────────

def load_feed(path):
    import csv as _csv
    for enc in ('utf-8', 'utf-8-sig', 'cp1252', 'latin-1'):
        try:
            with open(path, 'r', encoding=enc, newline='') as f:
                sample = f.read(4096)
            dialect = _csv.Sniffer().sniff(sample, delimiters='\t,|')
            sep = dialect.delimiter
            df = pd.read_csv(path, sep=sep, dtype=str, encoding=enc)
            df.columns = [c.strip().replace(' ', '_') for c in df.columns]
            return df
        except (UnicodeDecodeError, _csv.Error, pd.errors.ParserError):
            continue
    raise ValueError(f'Could not decode feed file: {path}')


# ── Vertical detection ─────────────────────────────────────────────────────────

def detect_vertical(df):
    cat_ids = set(df.get('google_product_category', pd.Series()).dropna().astype(str).unique())
    for vertical, ids in VERTICAL_CATEGORY_IDS.items():
        if cat_ids & ids:
            return vertical
    sample_text = ' '.join([
        ' '.join(df.get('product_type', pd.Series()).dropna().astype(str).tolist()),
        ' '.join(df.get('title', pd.Series()).dropna().astype(str).head(50).tolist()),
    ]).lower()
    for vertical, keywords in VERTICAL_KEYWORDS.items():
        if any(k in sample_text for k in keywords):
            return vertical
    return 'general'


def get_attr_sets(vertical):
    required = list(REQUIRED_UNIVERSAL) + REQUIRED_EXTRA.get(vertical, REQUIRED_EXTRA['general'])
    seen = set()
    required = [x for x in required if not (x in seen or seen.add(x))]
    recommended = list(RECOMMENDED_CORE) + RECOMMENDED_EXTRA.get(vertical, [])
    seen_r = set(required)
    recommended = [x for x in recommended if x not in seen_r and not (x in seen_r or seen_r.add(x))]
    return {'required': required, 'recommended': recommended}


# ── Audit logic ────────────────────────────────────────────────────────────────

def check_attr(df, attr):
    if attr not in df.columns:
        return 'NOT IN FEED', len(df), 100
    missing = int(df[attr].isna().sum() + (df[attr].eq('')).sum())
    pct = round(missing / len(df) * 100)
    return ('OK' if missing == 0 else 'INCOMPLETE'), missing, pct


def audit_attributes(df, vertical, skip_attrs=None):
    attr_sets = get_attr_sets(vertical)
    skip = skip_attrs or set()
    required_set    = set(attr_sets['required'])
    recommended_set = set(attr_sets['recommended'])
    rows = []
    for attr in ALL_ATTRIBUTES:
        if attr in required_set:
            # Required attrs are never skipped — always audited
            status, missing, pct = check_attr(df, attr)
            rows.append({'Attribute': attr, 'Type': 'Required',
                         'Status': status, 'Missing': missing, 'Missing %': pct})
        elif attr in recommended_set:
            if attr in skip:
                # Conditional attr whose signal wasn't found — demote to Not Recommended
                in_feed = attr in df.columns
                rows.append({'Attribute': attr, 'Type': 'Not Recommended',
                             'Status': 'Present' if in_feed else 'Not in feed',
                             'Missing': '—', 'Missing %': '—'})
            else:
                status, missing, pct = check_attr(df, attr)
                rows.append({'Attribute': attr, 'Type': 'Recommended',
                             'Status': status, 'Missing': missing, 'Missing %': pct})
        else:
            # Not in the vertical's required or recommended set — list for completeness
            in_feed = attr in df.columns
            rows.append({'Attribute': attr, 'Type': 'Not Recommended',
                         'Status': 'Present' if in_feed else 'Not in feed',
                         'Missing': '—', 'Missing %': '—'})
    return pd.DataFrame(rows)


def build_summary(df, attr_df, brand, vertical):
    total = len(df)
    req = attr_df[attr_df['Type'] == 'Required']
    rec = attr_df[attr_df['Type'] == 'Recommended']
    ok_req    = int((req['Status'] == 'OK').sum())
    total_req = len(req)
    any_missing = pd.Series([False] * total, index=df.index)
    for attr in req['Attribute'].tolist():
        if attr in df.columns:
            any_missing = any_missing | df[attr].isna() | df[attr].eq('')
        else:
            any_missing = pd.Series([True] * total, index=df.index)
    products_with_issues = int(any_missing.sum())
    present_rec = int((rec['Status'] != 'NOT IN FEED').sum())
    total_rec   = len(rec)
    return {
        'brand': brand,
        'vertical': vertical,
        'total_products': total,
        'products_with_issues': products_with_issues,
        'required_ok': ok_req,
        'required_total': total_req,
        'recommended_present': present_rec,
        'recommended_total': total_rec,
    }


def _feed_has_bundles(df):
    """True if feed appears to contain bundle products."""
    keywords = ['bundle', 'kit', 'set', 'combo', 'collection', 'pack']
    text = ' '.join([
        ' '.join(df.get('title', pd.Series()).dropna().astype(str).tolist()),
        ' '.join(df.get('product_type', pd.Series()).dropna().astype(str).tolist()),
    ]).lower()
    return any(k in text for k in keywords)


def _bundle_product_mask(df):
    """Boolean mask of rows that appear to be bundle products."""
    keywords = ['bundle', 'kit', 'set', 'combo', 'collection', 'pack']
    titles = df.get('title', pd.Series(dtype=str)).fillna('').str.lower()
    types  = df.get('product_type', pd.Series(dtype=str)).fillna('').str.lower()
    combined = titles + ' ' + types
    return combined.apply(lambda t: any(k in t for k in keywords))


def _feed_has_multipacks(df):
    """True if feed appears to contain multipack products."""
    keywords = ['2-pack', '3-pack', '4-pack', '6-pack', 'multipack', 'multi-pack', 'pack of']
    text = ' '.join(df.get('title', pd.Series()).dropna().astype(str).tolist()).lower()
    return any(k in text for k in keywords)


def _multipack_product_mask(df):
    """Boolean mask of rows that appear to be multipack products."""
    keywords = ['2-pack', '3-pack', '4-pack', '6-pack', 'multipack', 'multi-pack', 'pack of']
    titles = df.get('title', pd.Series(dtype=str)).fillna('').str.lower()
    return titles.apply(lambda t: any(k in t for k in keywords))


def _feed_has_gender_targeting(df):
    """True if products appear to target a specific gender."""
    keywords = ["women's", "men's", 'women', 'mens', 'womens', 'female', 'male',
                'his ', 'hers', 'ladies', 'girls', 'boys', 'for women', 'for men']
    text = ' '.join([
        ' '.join(df.get('title', pd.Series()).dropna().astype(str).tolist()),
        ' '.join(df.get('description', pd.Series()).dropna().astype(str).head(50).tolist()),
        ' '.join(df.get('product_type', pd.Series()).dropna().astype(str).tolist()),
    ]).lower()
    return any(k in text for k in keywords)


def _feed_has_color_variants(df):
    """True if product titles mention colours — indicates color is a purchase filter."""
    colors = ['black', 'white', 'red', 'blue', 'green', 'yellow', 'pink', 'purple',
              'orange', 'grey', 'gray', 'brown', 'navy', 'beige', 'gold', 'silver',
              'rose', 'teal', 'coral', 'ivory', 'color', 'colour']
    text = ' '.join(df.get('title', pd.Series()).dropna().astype(str).tolist()).lower()
    return any(c in text for c in colors)


def _feed_has_size_variants(df):
    """True if products have size/capacity/dimension variants."""
    import re
    text = ' '.join(df.get('title', pd.Series()).dropna().astype(str).tolist()).lower()
    size_keywords = ['small', 'medium', 'large', ' xl', ' xs', ' xxl', ' sm ',
                     'size', ' oz', ' ml', ' mg', 'gram', 'liter', 'litre',
                     'fl oz', 'count', 'capsule', 'tablet', 'serving']
    return (any(k in text for k in size_keywords) or
            bool(re.search(r'\d+\s*(oz|ml|mg|g|kg|lb|cm|mm|in|ct|cap|tab)\b', text)))


def _feed_has_material_variants(df):
    """True if product titles mention materials — indicates material is a purchase filter."""
    materials = ['cotton', 'polyester', 'nylon', 'leather', 'stainless', 'steel',
                 'aluminum', 'aluminium', 'wood', 'bamboo', 'plastic', 'rubber',
                 'silicone', 'ceramic', 'glass', 'canvas', 'denim', 'linen',
                 'wool', 'fleece', 'mesh', 'foam', 'fabric']
    text = ' '.join(df.get('title', pd.Series()).dropna().astype(str).tolist()).lower()
    return any(m in text for m in materials)


def _feed_has_pattern_variants(df):
    """True if product titles mention patterns — indicates pattern is a purchase filter."""
    patterns = ['striped', 'stripes', 'floral', 'plaid', 'checkered', 'camo',
                'camouflage', 'leopard', 'geometric', 'abstract', 'printed', 'pattern']
    text = ' '.join(df.get('title', pd.Series()).dropna().astype(str).tolist()).lower()
    return any(p in text for p in patterns)


def _feed_has_age_targeting(df):
    """True if products appear to target a specific age group."""
    keywords = ['baby', 'infant', 'toddler', 'kids', 'children', 'child',
                'adult', 'senior', 'elderly', 'teen', 'youth', 'newborn',
                'prenatal', 'postnatal', "children's", 'for kids', 'for babies']
    text = ' '.join([
        ' '.join(df.get('title', pd.Series()).dropna().astype(str).tolist()),
        ' '.join(df.get('description', pd.Series()).dropna().astype(str).head(50).tolist()),
    ]).lower()
    return any(k in text for k in keywords)


# Attrs that are only recommended when the feed contains signals that make them relevant.
# Each entry maps an attribute name to a detection function.
# If the function returns False → attribute is demoted to "Not Recommended" in the audit.
CONDITIONAL_ATTRS = {
    'is_bundle':  _feed_has_bundles,
    'multipack':  _feed_has_multipacks,
    'gender':     _feed_has_gender_targeting,
    'color':      _feed_has_color_variants,
    'size':       _feed_has_size_variants,
    'material':   _feed_has_material_variants,
    'pattern':    _feed_has_pattern_variants,
    'age_group':  _feed_has_age_targeting,
}


def build_issues(df, attr_df, vertical):
    """One row per specific issue, ranked by priority."""
    issues = []
    n = len(df)
    priority = 1

    # Pre-evaluate conditional attr relevance
    skip_attrs = {attr for attr, check_fn in CONDITIONAL_ATTRS.items() if not check_fn(df)}

    # These attrs have dedicated checks below — exclude from generic sections 10/11
    DEDICATED_ATTRS = {'is_bundle', 'multipack', 'product_type'}
    skip_attrs |= DEDICATED_ATTRS

    req_rows = attr_df[attr_df['Type'] == 'Required']
    rec_rows = attr_df[attr_df['Type'] == 'Recommended']

    # ── 1. Title issues (one row per issue type) ───────────────────────────────
    if 'title' in df.columns:
        lengths = df['title'].str.len().fillna(0)
        short_t = int((lengths < 25).sum())
        long_t  = int((lengths > 150).sum())
        dup_t   = int(df['title'].duplicated(keep=False).sum())

        if short_t > 0:
            issues.append({
                '#': priority,
                'Attribute': 'title',
                'Issue': f'{short_t} titles too short (<25 chars)',
                'Affected Products': short_t,
                '%': f'{round(short_t / n * 100)}%',
                'What to do': 'Expand titles — add brand, product type, key attribute (size, flavour, count). Target 70–150 chars. → Run Part 2 of this skill to auto-generate optimised titles.',
            })
            priority += 1

        if dup_t > 0:
            issues.append({
                '#': priority,
                'Attribute': 'title',
                'Issue': f'{dup_t} duplicate titles across variants',
                'Affected Products': dup_t,
                '%': f'{round(dup_t / n * 100)}%',
                'What to do': 'Add variant-specific attributes (size, colour, count, flavour) to make each title unique. → Run Part 2 of this skill to auto-generate optimised titles.',
            })
            priority += 1

        if long_t > 0:
            issues.append({
                '#': priority,
                'Attribute': 'title',
                'Issue': f'{long_t} titles too long (>150 chars)',
                'Affected Products': long_t,
                '%': f'{round(long_t / n * 100)}%',
                'What to do': 'Trim to 150 chars max. Front-load the most important keywords — Google truncates from the right. → Run Part 2 of this skill to auto-generate optimised titles.',
            })
            priority += 1

    # ── 2. Description issues (one row per issue type) ─────────────────────────
    if 'description' in df.columns:
        lengths = df['description'].str.len().fillna(0)
        short_d = int((lengths < 100).sum())
        long_d  = int((lengths > 5000).sum())
        dup_d   = int(df['description'].duplicated(keep=False).sum())

        if short_d > 0:
            issues.append({
                '#': priority,
                'Attribute': 'description',
                'Issue': f'{short_d} descriptions too short (<100 chars)',
                'Affected Products': short_d,
                '%': f'{round(short_d / n * 100)}%',
                'What to do': 'Write unique descriptions listing key attributes and specs. No marketing copy. Target 500–2000 chars. → Run Part 3 of this skill to auto-generate optimised descriptions.',
            })
            priority += 1

        if dup_d > 0:
            issues.append({
                '#': priority,
                'Attribute': 'description',
                'Issue': f'{dup_d} duplicate descriptions',
                'Affected Products': dup_d,
                '%': f'{round(dup_d / n * 100)}%',
                'What to do': 'Every description must be unique. Tailor each one to the specific product variant. → Run Part 3 of this skill to auto-generate optimised descriptions.',
            })
            priority += 1

        if long_d > 0:
            issues.append({
                '#': priority,
                'Attribute': 'description',
                'Issue': f'{long_d} descriptions too long (>5000 chars)',
                'Affected Products': long_d,
                '%': f'{round(long_d / n * 100)}%',
                'What to do': 'Trim to under 5000 chars. Keep the most relevant product information at the top. → Run Part 3 of this skill to auto-generate optimised descriptions.',
            })
            priority += 1

    # ── 3. Required attrs NOT IN FEED ─────────────────────────────────────────
    for _, row in req_rows[req_rows['Status'] == 'NOT IN FEED'].iterrows():
        attr = row['Attribute']
        issues.append({
            '#': priority,
            'Attribute': attr,
            'Issue': 'Missing from feed — column does not exist',
            'Affected Products': n,
            '%': '100%',
            'What to do': _what_to_do(attr),
        })
        priority += 1

    # ── 4. Required attrs INCOMPLETE ──────────────────────────────────────────
    for _, row in req_rows[req_rows['Status'] == 'INCOMPLETE'].iterrows():
        attr = row['Attribute']
        count = int(row['Missing'])
        pct   = int(row['Missing %'])
        issues.append({
            '#': priority,
            'Attribute': attr,
            'Issue': f'Blank on {count} products — required attribute',
            'Affected Products': count,
            '%': f'{pct}%',
            'What to do': _what_to_do(attr),
        })
        priority += 1

    # ── 5. google_product_category too broad ──────────────────────────────────
    if 'google_product_category' in df.columns:
        valid_cats = df['google_product_category'].dropna()
        valid_cats = valid_cats[valid_cats != '']
        broad_mask = valid_cats.apply(
            lambda v: str(v) in TAXONOMY and decode_category(str(v)).count('>') == 0
        )
        broad_count = int(broad_mask.sum())
        if broad_count > 0:
            pct = round(broad_count / n * 100)
            issues.append({
                '#': priority,
                'Attribute': 'google_product_category',
                'Issue': f'Too broad — top-level category on {broad_count} products',
                'Affected Products': broad_count,
                '%': f'{pct}%',
                'What to do': "Replace with the most specific subcategory from Google's taxonomy (e.g. Health Care > Vitamins & Supplements, not Health & Beauty).",
            })
            priority += 1

    # ── 6. product_type quality checks (per Google's spec) ───────────────────
    if 'product_type' in df.columns:
        populated = df['product_type'].dropna()
        populated = populated[populated != '']

        # 6a. No hierarchy (missing ">" separator)
        flat_mask = ~populated.str.contains('>', na=False)
        flat_count = int(flat_mask.sum())
        if flat_count > 0:
            pct = round(flat_count / n * 100)
            issues.append({
                '#': priority,
                'Attribute': 'product_type',
                'Issue': f'Not structured as a hierarchy on {flat_count} products — missing ">" separator',
                'Affected Products': flat_count,
                '%': f'{pct}%',
                'What to do': 'Structure your product_type as a hierarchy using ">" as the separator, going from broad to specific (e.g. Supplements > Probiotics > Women\'s Probiotics > 50 Billion CFU Capsules). Google reads this attribute to understand semantically what you\'re selling — a flat single value gives it nothing to work with.',
            })
            priority += 1

        # 6b. Too shallow — only 2 levels (exactly one ">")
        two_level_mask = populated.str.contains('>', na=False) & (populated.str.count(r'>') == 1)
        two_level_count = int(two_level_mask.sum())
        if two_level_count > 0:
            pct = round(two_level_count / n * 100)
            issues.append({
                '#': priority,
                'Attribute': 'product_type',
                'Issue': f'Only 2 levels deep on {two_level_count} products — Google recommends 3+ levels',
                'Affected Products': two_level_count,
                '%': f'{pct}%',
                'What to do': 'Add more specificity — go deeper than 2 levels. Google uses product_type to semantically understand what you\'re selling, so the more descriptive the path, the better. Good example: Supplements > Probiotics > Women\'s Probiotics > 50 Billion CFU Capsules. You can also incorporate key attributes at deeper levels (e.g. format, strength, target audience).',
            })
            priority += 1

        # 6c. Exceeds 750-character limit (Google's hard limit)
        over_limit_mask = populated.str.len() > 750
        over_limit_count = int(over_limit_mask.sum())
        if over_limit_count > 0:
            pct = round(over_limit_count / n * 100)
            issues.append({
                '#': priority,
                'Attribute': 'product_type',
                'Issue': f'Exceeds 750-character limit on {over_limit_count} products',
                'Affected Products': over_limit_count,
                '%': f'{pct}%',
                'What to do': 'Shorten the product_type value to 750 characters or fewer. Google will truncate or reject values that exceed this limit.',
            })
            priority += 1

        # 6d. Contains HTML tags
        html_mask = populated.str.contains(r'<[^>]+>', na=False, regex=True)
        html_count = int(html_mask.sum())
        if html_count > 0:
            pct = round(html_count / n * 100)
            issues.append({
                '#': priority,
                'Attribute': 'product_type',
                'Issue': f'Contains HTML tags on {html_count} products — plain text required',
                'Affected Products': html_count,
                '%': f'{pct}%',
                'What to do': 'Remove all HTML tags from product_type values. Use plain text only.',
            })
            priority += 1

    # 6e. product_type blank or missing (handled here, not in generic sections 10/11)
    pt_status, pt_missing, pt_pct = check_attr(df, 'product_type')
    if pt_status == 'NOT IN FEED':
        issues.append({
            '#': priority,
            'Attribute': 'product_type',
            'Issue': 'Not in feed — recommended attribute',
            'Affected Products': n,
            '%': '100%',
            'What to do': _what_to_do('product_type'),
        })
        priority += 1
    elif pt_status == 'INCOMPLETE' and pt_missing > 0:
        issues.append({
            '#': priority,
            'Attribute': 'product_type',
            'Issue': f'Blank on {pt_missing} products ({pt_pct}%) — recommended attribute',
            'Affected Products': pt_missing,
            '%': f'{pt_pct}%',
            'What to do': _what_to_do('product_type'),
        })
        priority += 1

    # ── 7. Invalid values on key required attrs ────────────────────────────────
    for attr, valid_set in VALID_VALUES.items():
        if attr in df.columns:
            populated = df[attr].dropna()
            populated = populated[populated != '']
            invalid_mask = ~populated.str.lower().isin(valid_set)
            count = int(invalid_mask.sum())
            if count > 0:
                pct = round(count / n * 100)
                issues.append({
                    '#': priority,
                    'Attribute': attr,
                    'Issue': f'Invalid value on {count} products',
                    'Affected Products': count,
                    '%': f'{pct}%',
                    'What to do': f'Accepted values: {" | ".join(sorted(valid_set))}.',
                })
                priority += 1

    # ── 8. is_bundle missing on products that look like bundles ──────────────
    bundle_mask = _bundle_product_mask(df)
    bundle_count = int(bundle_mask.sum())
    if bundle_count > 0:
        if 'is_bundle' in df.columns:
            missing_bundle = bundle_mask & (df['is_bundle'].isna() | df['is_bundle'].eq('') | ~df['is_bundle'].str.lower().eq('yes'))
        else:
            missing_bundle = bundle_mask
        missing_count = int(missing_bundle.sum())
        if missing_count > 0:
            pct = round(missing_count / n * 100)
            issues.append({
                '#': priority,
                'Attribute': 'is_bundle',
                'Issue': f'Missing or not set to "yes" on {missing_count} likely bundle products',
                'Affected Products': missing_count,
                '%': f'{pct}%',
                'What to do': 'Set is_bundle to "yes" for products that contain multiple distinct items sold together (e.g. a gift set or kit). This helps Google correctly classify the product and can unlock bundle-specific treatment.',
            })
            priority += 1

    # ── 9. multipack missing on products that look like multipacks ────────────
    multipack_mask = _multipack_product_mask(df)
    multipack_count = int(multipack_mask.sum())
    if multipack_count > 0:
        if 'multipack' in df.columns:
            missing_mp = multipack_mask & (df['multipack'].isna() | df['multipack'].eq(''))
        else:
            missing_mp = multipack_mask
        missing_mp_count = int(missing_mp.sum())
        if missing_mp_count > 0:
            pct = round(missing_mp_count / n * 100)
            issues.append({
                '#': priority,
                'Attribute': 'multipack',
                'Issue': f'Missing on {missing_mp_count} likely multipack products',
                'Affected Products': missing_mp_count,
                '%': f'{pct}%',
                'What to do': 'Set multipack to the number of identical units in the pack (e.g. 3 for a 3-pack). Required when the product is a merchant-defined multipack.',
            })
            priority += 1

    # ── 10. Recommended attrs NOT IN FEED ─────────────────────────────────────
    for _, row in rec_rows[rec_rows['Status'] == 'NOT IN FEED'].iterrows():
        attr = row['Attribute']
        if attr in skip_attrs:
            continue
        issues.append({
            '#': priority,
            'Attribute': attr,
            'Issue': 'Not in feed — recommended attribute',
            'Affected Products': n,
            '%': '100%',
            'What to do': _what_to_do(attr),
        })
        priority += 1

    # ── 11. Recommended attrs INCOMPLETE (any missing) ────────────────────────
    for _, row in rec_rows[rec_rows['Status'] == 'INCOMPLETE'].iterrows():
        if int(row['Missing']) > 0:
            attr  = row['Attribute']
            if attr in skip_attrs:
                continue
            count = int(row['Missing'])
            pct   = int(row['Missing %'])
            issues.append({
                '#': priority,
                'Attribute': attr,
                'Issue': f'Blank on {count} products ({pct}%) — recommended attribute',
                'Affected Products': count,
                '%': f'{pct}%',
                'What to do': _what_to_do(attr),
            })
            priority += 1

    if not issues:
        return pd.DataFrame(columns=['#', 'Attribute', 'Issue', 'Affected Products', '%', 'What to do'])
    return pd.DataFrame(issues)


def build_product_issues(df, attr_df):
    """One row per product × issue. Columns: Product ID | Title | Attribute | Issue | Current Value.
    For 'not in feed' issues (no per-product variation), a single summary row is added instead of
    one row per product — avoids table bloat on large feeds while keeping the filter useful."""
    import re as _re

    rows = []
    id_col    = 'id' if 'id' in df.columns else df.columns[0]
    title_col = 'title' if 'title' in df.columns else None
    n = len(df)

    def get_id(row):
        v = row.get(id_col, None)
        return str(v) if v is not None and not pd.isna(v) else '(blank)'

    def get_title(row):
        if title_col:
            v = row.get(title_col, None)
            if v is not None and not pd.isna(v) and str(v).strip():
                t = str(v)
                return (t[:77] + '…') if len(t) > 80 else t
        return '(no title)'

    def add(row, attr, issue, current_val):
        rows.append({
            'Product ID':    get_id(row),
            'Title':         get_title(row),
            'Attribute':     attr,
            'Issue':         issue,
            'Current Value': str(current_val)[:200] if current_val is not None else '',
        })

    def add_summary(attr, issue, current_val, count=None):
        label = f'(all {count} products)' if count is not None else '(all products)'
        rows.append({
            'Product ID':    label,
            'Title':         '—',
            'Attribute':     attr,
            'Issue':         issue,
            'Current Value': str(current_val),
        })

    req_rows = attr_df[attr_df['Type'] == 'Required']
    rec_rows = attr_df[attr_df['Type'] == 'Recommended']

    # ── 1. Title quality ───────────────────────────────────────────────────────
    if 'title' in df.columns:
        dup_titles = set(
            df['title'].dropna()[df['title'].dropna().duplicated(keep=False)].tolist()
        )
        for _, row in df.iterrows():
            val = row.get('title', None)
            if pd.isna(val) or str(val).strip() == '':
                continue
            sval = str(val)
            length = len(sval)
            display = (sval[:77] + '…') if len(sval) > 80 else sval
            if length < 25:
                add(row, 'title', 'Too short (<25 chars)', display)
            if length > 150:
                add(row, 'title', 'Too long (>150 chars)', f'{length} chars — {display}')
            if sval in dup_titles:
                add(row, 'title', 'Duplicate title', display)

    # ── 2. Description quality ─────────────────────────────────────────────────
    if 'description' in df.columns:
        dup_descs = set(
            df['description'].dropna()[df['description'].dropna().duplicated(keep=False)].tolist()
        )
        for _, row in df.iterrows():
            val = row.get('description', None)
            if pd.isna(val) or str(val).strip() == '':
                continue
            sval = str(val)
            length = len(sval)
            if length < 100:
                add(row, 'description', 'Too short (<100 chars)', f'{length} chars')
            if length > 5000:
                add(row, 'description', 'Too long (>5000 chars)', f'{length} chars')
            if sval in dup_descs:
                add(row, 'description', 'Duplicate description', f'{length} chars')

    # ── 3. Required attrs NOT IN FEED — summary row (no per-product detail) ───
    for _, arow in req_rows[req_rows['Status'] == 'NOT IN FEED'].iterrows():
        add_summary(arow['Attribute'], 'Required attribute not in feed', '(not in feed)', n)

    # ── 4. Required attrs blank — one row per affected product ─────────────────
    for _, arow in req_rows[req_rows['Status'] == 'INCOMPLETE'].iterrows():
        attr = arow['Attribute']
        if attr not in df.columns:
            continue
        for _, row in df[df[attr].isna() | df[attr].eq('')].iterrows():
            add(row, attr, 'Required attribute blank', '(blank)')

    # ── 5. google_product_category too broad ──────────────────────────────────
    if 'google_product_category' in df.columns:
        for _, row in df.iterrows():
            val = row.get('google_product_category', None)
            if pd.isna(val) or str(val).strip() == '':
                continue
            sval = str(val).strip()
            if sval in TAXONOMY and decode_category(sval).count('>') == 0:
                add(row, 'google_product_category', 'Too broad (top-level category)', decode_category(sval))

    # ── 6. product_type quality ────────────────────────────────────────────────
    if 'product_type' in df.columns:
        for _, row in df.iterrows():
            val = row.get('product_type', None)
            if pd.isna(val) or str(val).strip() == '':
                add(row, 'product_type', 'Blank — recommended attribute', '(blank)')
                continue
            sval = str(val).strip()
            display = (sval[:77] + '…') if len(sval) > 80 else sval
            if '>' not in sval:
                add(row, 'product_type', 'Not structured as a hierarchy (missing ">")', display)
            elif sval.count('>') == 1:
                add(row, 'product_type', 'Only 2 levels deep (recommend 3+)', display)
            if len(sval) > 750:
                add(row, 'product_type', 'Exceeds 750-character limit', f'{len(sval)} chars')
            if _re.search(r'<[^>]+>', sval):
                add(row, 'product_type', 'Contains HTML tags', display)
    else:
        add_summary('product_type', 'Recommended attribute not in feed', '(not in feed)', n)

    # ── 7. Invalid values ──────────────────────────────────────────────────────
    for attr, valid_set in VALID_VALUES.items():
        if attr in df.columns:
            for _, row in df.iterrows():
                val = row.get(attr, None)
                if pd.isna(val) or str(val).strip() == '':
                    continue
                if str(val).strip().lower() not in valid_set:
                    add(row, attr, 'Invalid value', str(val).strip())

    # ── 8. is_bundle missing on bundle products ────────────────────────────────
    bundle_mask = _bundle_product_mask(df)
    for idx in df.index[bundle_mask]:
        row = df.loc[idx]
        if 'is_bundle' in df.columns:
            val = row.get('is_bundle', None)
            already_set = (val is not None and not pd.isna(val) and str(val).strip().lower() == 'yes')
        else:
            val = None
            already_set = False
        if not already_set:
            current = str(val).strip() if (val is not None and not pd.isna(val) and str(val).strip()) else '(blank)'
            add(row, 'is_bundle', 'Missing or not "yes" on likely bundle product', current)

    # ── 9. multipack missing on multipack products ─────────────────────────────
    multipack_mask = _multipack_product_mask(df)
    for idx in df.index[multipack_mask]:
        row = df.loc[idx]
        if 'multipack' in df.columns:
            val = row.get('multipack', None)
            is_blank = val is None or pd.isna(val) or str(val).strip() == ''
        else:
            is_blank = True
        if is_blank:
            add(row, 'multipack', 'Missing on likely multipack product', '(blank)')

    # ── 10. Recommended attrs NOT IN FEED — summary row ───────────────────────
    DEDICATED_ATTRS = {'is_bundle', 'multipack', 'product_type'}
    skip_attrs = {attr for attr, check_fn in CONDITIONAL_ATTRS.items() if not check_fn(df)}
    skip_attrs |= DEDICATED_ATTRS
    for _, arow in rec_rows[rec_rows['Status'] == 'NOT IN FEED'].iterrows():
        attr = arow['Attribute']
        if attr in skip_attrs:
            continue
        add_summary(attr, 'Recommended attribute not in feed', '(not in feed)', n)

    # ── 11. Recommended attrs blank — one row per affected product ─────────────
    for _, arow in rec_rows[rec_rows['Status'] == 'INCOMPLETE'].iterrows():
        attr = arow['Attribute']
        if attr in skip_attrs or attr not in df.columns:
            continue
        if int(arow['Missing']) == 0:
            continue
        for _, row in df[df[attr].isna() | df[attr].eq('')].iterrows():
            add(row, attr, 'Recommended attribute blank', '(blank)')

    if not rows:
        return pd.DataFrame(columns=['Product ID', 'Title', 'Attribute', 'Issue', 'Current Value'])

    result = pd.DataFrame(rows)
    result = result.sort_values(['Attribute', 'Issue']).reset_index(drop=True)
    return result


# ── Excel sheets ───────────────────────────────────────────────────────────────

def write_sheet_readme(wb, summary):
    ws = wb.create_sheet('README')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 28
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18

    today = __import__('datetime').date.today().strftime('%B %d, %Y')
    brand = summary['brand']

    def sec_header(row, text):
        ws.merge_cells(f'A{row}:F{row}')
        c = ws[f'A{row}']
        c.value = text
        c.font  = Font(name='Calibri', bold=True, size=10, color=C_WHITE)
        c.fill  = fill(C_ACCENT)
        c.alignment = left()
        c.border = border
        ws.row_dimensions[row].height = 22

    def body(row, text, indent=False):
        ws.merge_cells(f'A{row}:F{row}')
        c = ws[f'A{row}']
        c.value = ('    ' + text) if indent else text
        c.font  = Font(name='Calibri', size=10, color='FF2B303B')
        c.fill  = fill('FFF8F9FF')
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 36

    def tab_row(row, name, desc):
        ws.row_dimensions[row].height = 36
        a = ws.cell(row=row, column=1, value=name)
        a.font  = Font(name='Calibri', bold=True, size=10, color='FF3664F4')
        a.fill  = fill('FFF0F4FF')
        a.alignment = left()
        a.border = border
        ws.merge_cells(f'B{row}:F{row}')
        b = ws[f'B{row}']
        b.value = desc
        b.font  = Font(name='Calibri', size=10, color='FF2B303B')
        b.fill  = fill('FFF8F9FF')
        b.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        b.border = border

    def swatch_row(row, bg, text):
        ws.row_dimensions[row].height = 30
        s = ws.cell(row=row, column=1, value='')
        s.fill   = fill(bg)
        s.border = border
        ws.merge_cells(f'B{row}:F{row}')
        t = ws[f'B{row}']
        t.value = text
        t.font  = Font(name='Calibri', size=10, color='FF2B303B')
        t.fill  = fill('FFF8F9FF')
        t.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        t.border = border

    def step_row(row, number, text):
        ws.row_dimensions[row].height = 30
        n = ws.cell(row=row, column=1, value=f'Step {number}')
        n.font  = Font(name='Calibri', bold=True, size=10, color=C_WHITE)
        n.fill  = fill(C_SUB)
        n.alignment = center()
        n.border = border
        ws.merge_cells(f'B{row}:F{row}')
        t = ws[f'B{row}']
        t.value = text
        t.font  = Font(name='Calibri', size=10, color='FF2B303B')
        t.fill  = fill('FFF8F9FF')
        t.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        t.border = border

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:F1')
    ws['A1'].value = 'PRODUCT FEED AUDIT — HOW TO USE THIS REPORT'
    ws['A1'].font  = hfont(size=14)
    ws['A1'].fill  = fill(C_HEADER)
    ws['A1'].alignment = center()
    ws.row_dimensions[1].height = 40

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    ws.merge_cells('A2:F2')
    ws['A2'].value = f'{brand}  ·  Generated {today}'
    ws['A2'].font  = Font(name='Calibri', bold=False, size=10, color='FFB0B8CC')
    ws['A2'].fill  = fill(C_SUB)
    ws['A2'].alignment = center()
    ws.row_dimensions[2].height = 22

    # ── Row 3: Spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 10

    # ── What is this file? ────────────────────────────────────────────────────
    sec_header(4, 'WHAT IS THIS FILE?')
    body(5, (
        f'This report is an automated audit of {brand}\'s Google Shopping product feed. '
        'It scans every product and attribute against Google\'s feed requirements and best practices, '
        'then surfaces issues ranked by impact on campaign performance and product approval.'
    ))
    ws.row_dimensions[5].height = 52
    ws.row_dimensions[6].height = 8  # spacer

    # ── Tabs in this report ───────────────────────────────────────────────────
    sec_header(7, 'TABS IN THIS REPORT')
    tab_row(8,  '1 — Issues Summary',
                'A prioritised list of every issue found, ranked by impact. Start here. '
                'Each row shows the affected attribute, the number of products impacted, '
                'and a clear action to fix it.')
    tab_row(9,  '2 — Attribute Audit',
                'A full checklist of all Google Shopping attributes — required, recommended, '
                'and informational — with status (OK / Incomplete / Not in feed) for each one.')
    tab_row(10, '3 — All Products',
                'The complete product catalog with cell-level colour coding so you can '
                'see exactly which products have issues in which fields.')
    tab_row(11, '4 — Product Issues',
                'A filterable table — one row per product × issue. Click the dropdown (▼) on '
                'the "Issue" column to instantly see every product with a specific problem. '
                'Filter to "Too short" to find all short-title products, or "Required attribute blank" '
                'to find missing fields. Native Excel/Sheets filtering — no macros required.')
    ws.row_dimensions[12].height = 10  # spacer

    # ── Colour guide ──────────────────────────────────────────────────────────
    sec_header(13, 'COLOUR GUIDE')
    swatch_row(14, 'FFFFEBEE', '🔴  Red — Required attribute is missing or blank. These products risk disapproval in Google Merchant Center. Fix these first.')
    swatch_row(15, 'FFFFF8E1', '🟡  Yellow — Quality issue or recommended attribute missing. Products will still serve but performance may suffer.')
    swatch_row(16, C_GREY_FILL, '⬜  Grey — Informational column. Not audited (Google does not require or recommend this attribute for your vertical).')
    swatch_row(17, C_WHITE,     '✅  White — No issue detected. This cell is OK.')
    ws.row_dimensions[18].height = 10  # spacer

    # ── How to find problem products ──────────────────────────────────────────
    sec_header(19, 'HOW TO FIND PROBLEM PRODUCTS')
    body(20, 'QUICKEST — Product Issues tab: Go to the "4 — Product Issues" tab. Click the dropdown arrow (▼) on the "Issue" column header. Select the issue type you want to investigate. All affected products appear instantly — no scrolling needed.')
    ws.row_dimensions[20].height = 52

    body(21, 'ALTERNATIVE — All Products tab colour filtering (Microsoft Excel):')
    ws['A21'].font = Font(name='Calibri', bold=True, size=10, color='FF3664F4')
    ws.row_dimensions[21].height = 24
    step_row(22, 1, 'Click any cell inside the data table.')
    step_row(23, 2, 'Go to Data → Filter  (or press Ctrl + Shift + L)  to enable column filters.')
    step_row(24, 3, 'Click the dropdown arrow (▼) on any column header — e.g. "title" or "image_link".')
    step_row(25, 4, 'Choose  Filter by Color → Filter by Cell Color.')
    step_row(26, 5, 'Select the red or yellow fill to show only affected products. Clear the filter to see all products again.')

    ws.row_dimensions[27].height = 10  # spacer

    body(28, 'ALTERNATIVE — All Products tab colour filtering (Google Sheets):')
    ws['A28'].font = Font(name='Calibri', bold=True, size=10, color='FF3664F4')
    ws.row_dimensions[28].height = 24
    step_row(29, 1, 'Select all data (Ctrl + A), then go to Data → Create a filter.')
    step_row(30, 2, 'Click the filter icon on a column header.')
    step_row(31, 3, 'Choose  Filter by condition → Custom formula,  then enter  =ISBLANK(A2)  to find blanks in that column.')
    step_row(32, 4, 'Alternatively, sort the column by fill color using  Data → Sort range → Advanced sort options.')

    ws.row_dimensions[33].height = 16  # bottom padding


def write_sheet_summary(wb, summary, issues_df):
    ws = wb.create_sheet('Issues Summary')
    ws.sheet_view.showGridLines = False

    # Table has 6 columns (A–F): #, Attribute, Issue, Affected Products, %, What to do
    # All header merges must stay within A:F.

    # ── Title block ────────────────────────────────────────────────────────────
    ws.merge_cells('A1:F1')
    ws['A1'].value = 'PRODUCT FEED AUDIT — ISSUES SUMMARY'
    ws['A1'].font = hfont(size=14)
    ws['A1'].fill = fill(C_HEADER)
    ws['A1'].alignment = center()

    ws.merge_cells('A2:F2')
    vertical_label = VERTICAL_LABELS.get(summary['vertical'], summary['vertical'].title())
    ws['A2'].value = f'{summary["brand"]}  ·  {datetime.today().strftime("%d %b %Y")}  ·  {vertical_label}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='FF9BA4B5')
    ws['A2'].fill = fill(C_HEADER)
    ws['A2'].alignment = center()

    # ── Scorecard — 2×2 grid within A:F ────────────────────────────────────────
    # Row 4–5: left card (A:C) = Total Products, right card (D:F) = Products w/ Issues
    # Row 6–7: left card (A:C) = Required Attributes, right card (D:F) = Recommended Attributes
    ws.row_dimensions[5].height = 38
    ws.row_dimensions[7].height = 38

    req_color = ('FF43A047' if summary['required_ok'] == summary['required_total']
                 else ('FFFB8C00' if summary['required_ok'] >= summary['required_total'] * 0.7
                       else 'FFE53935'))
    rec_pct   = summary['recommended_present'] / summary['recommended_total'] if summary['recommended_total'] else 1
    rec_color = 'FF43A047' if rec_pct >= 0.8 else ('FFFB8C00' if rec_pct >= 0.5 else 'FFE53935')

    scorecard = [
        # (label_range, val_range, label, value, color)
        ('A4:C4', 'A5:C5', 'TOTAL PRODUCTS',        str(summary['total_products']),                                      'FF3664F4'),
        ('D4:F4', 'D5:F5', 'PRODUCTS W/ ISSUES',    str(summary['products_with_issues']),                                'FFE53935' if summary['products_with_issues'] > 0 else 'FF43A047'),
        ('A6:C6', 'A7:C7', 'REQUIRED ATTRIBUTES',   f'{summary["required_ok"]}/{summary["required_total"]}',             req_color),
        ('D6:F6', 'D7:F7', 'RECOMMENDED ATTRIBUTES', f'{summary["recommended_present"]}/{summary["recommended_total"]}', rec_color),
    ]

    card_border = Border(
        left=Side(style='thin', color='FFD0D7E8'),
        right=Side(style='thin', color='FFD0D7E8'),
        top=Side(style='thin', color='FFD0D7E8'),
        bottom=Side(style='thin', color='FFD0D7E8'),
    )
    card_fill = fill('FFF0F4FF')

    for label_range, val_range, label, value, color in scorecard:
        ws.merge_cells(label_range)
        lc = ws[label_range.split(':')[0]]
        lc.value = label
        lc.font = Font(name='Calibri', bold=True, size=9, color='FF737B8C')
        lc.fill = card_fill
        lc.alignment = center()
        lc.border = card_border

        ws.merge_cells(val_range)
        vc = ws[val_range.split(':')[0]]
        vc.value = value
        vc.font = Font(name='Calibri', bold=True, size=22, color=color)
        vc.fill = card_fill
        vc.alignment = center()
        vc.border = card_border

    # ── Color legend (row 8) ───────────────────────────────────────────────────
    ws.row_dimensions[8].height = 20
    legend = [
        ('A8:B8', 'FFFFEBEE', C_RED_TEXT,   'Red — Required attribute missing or blank'),
        ('C8:D8', 'FFFFF8E1', C_AMBER_TEXT,  'Yellow — Quality issue or recommended attribute missing'),
        ('E8:F8', 'FFFFFFFF', 'FF737B8C',    'White — Informational / lower priority'),
    ]
    leg_border = Border(
        left=Side(style='thin', color='FFD0D7E8'),
        right=Side(style='thin', color='FFD0D7E8'),
        top=Side(style='thin', color='FFD0D7E8'),
        bottom=Side(style='thin', color='FFD0D7E8'),
    )
    for cell_range, bg, txt_color, label in legend:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(':')[0]]
        cell.value = label
        cell.fill = fill(bg)
        cell.font = Font(name='Calibri', size=9, bold=False, color=txt_color)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = leg_border

    # ── Issues table header ────────────────────────────────────────────────────
    ws.merge_cells('A9:F9')
    ws['A9'].value = f'ISSUES  ({len(issues_df)} found)'
    ws['A9'].font = hfont(size=11)
    ws['A9'].fill = fill(C_SUB)
    ws['A9'].alignment = left()

    headers = ['#', 'Attribute', 'Issue', 'Affected Products', '%', 'What to do']
    write_header_row(ws, 10, headers, bg=C_ACCENT)

    # Priority bands: rows 1–(last required issue) = critical; then warnings; then info
    # We determine band by issue content
    for r, (_, row) in enumerate(issues_df.iterrows(), 11):
        issue_text = str(row['Issue']).lower()
        attr       = str(row['Attribute'])

        is_critical = (
            'missing from feed' in issue_text or
            ('blank' in issue_text and attr in set(REQUIRED_UNIVERSAL))
        )
        is_warning = (
            'too broad' in issue_text or
            'invalid' in issue_text or
            'duplicate' in issue_text or
            'too short' in issue_text or
            'too long' in issue_text or
            'not structured as a hierarchy' in issue_text or
            'only 2 levels deep' in issue_text or
            'exceeds 750' in issue_text or
            'html tags' in issue_text or
            'not in feed' in issue_text or
            ('blank' in issue_text and attr not in set(REQUIRED_UNIVERSAL))
        )

        if is_critical:
            row_bg = 'FFFFEBEE'   # very light red
            num_color = C_RED_TEXT
        elif is_warning:
            row_bg = 'FFFFF8E1'   # very light amber
            num_color = C_AMBER_TEXT
        else:
            row_bg = 'FFFFFFFF'
            num_color = 'FF737B8C'

        ws.row_dimensions[r].height = 30

        for c, key in enumerate(['#', 'Attribute', 'Issue', 'Affected Products', '%', 'What to do'], 1):
            val  = row[key]
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill(row_bg)
            cell.border = border
            cell.alignment = left(wrap=(c == 6))

            if c == 1:
                cell.font = Font(name='Calibri', bold=True, size=10, color=num_color)
                cell.alignment = center()
            elif c == 2:
                cell.font = Font(name='Calibri', bold=True, size=10, color='FF2B303B')
            elif c in (4, 5):
                cell.font = Font(name='Calibri', size=10, color='FF2B303B')
                cell.alignment = center()
            else:
                cell.font = Font(name='Calibri', size=10, color='FF2B303B')

    set_col_widths(ws, [4, 22, 46, 18, 7, 65])
    ws.freeze_panes = 'A9'


def write_sheet_attributes(wb, attr_df):
    ws = wb.create_sheet('Attribute Audit')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:E1')
    ws['A1'].value = 'ATTRIBUTE AUDIT'
    ws['A1'].font = hfont(size=13)
    ws['A1'].fill = fill(C_HEADER)
    ws['A1'].alignment = center()

    write_header_row(ws, 2, ['Attribute', 'Type', 'Status', 'Missing (count)', 'Missing (%)'])

    status_styles = {
        'OK':          ('FF2E7D32', False),
        'INCOMPLETE':  ('FFE65100', True),
        'NOT IN FEED': ('FFC62828', True),
        'Present':     ('FF9BA4B5', False),
        'Not in feed': ('FF9BA4B5', False),
    }

    C_NOT_REC = 'FFF5F5F5'   # very light grey background for non-applicable rows

    for r, (_, row) in enumerate(attr_df.iterrows(), 3):
        attr_type = str(row['Type'])
        is_req        = attr_type == 'Required'
        is_not_rec    = attr_type == 'Not Recommended'

        if is_not_rec:
            bg = C_NOT_REC
        else:
            bg = C_LIGHT if r % 2 == 0 else C_WHITE

        color, bold = status_styles.get(str(row['Status']), ('FF2B303B', False))

        for c, val in enumerate([row['Attribute'], row['Type'], row['Status'],
                                  row['Missing'], row['Missing %']], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.fill   = fill(bg)

            if is_not_rec:
                # Muted grey text for non-applicable attributes
                cell.font      = Font(name='Calibri', size=10,
                                      bold=(c == 1), color='FFBBBFC8')
                cell.alignment = center() if c != 1 else left()
            elif c == 1:
                cell.font      = Font(name='Calibri', bold=True, size=10, color='FF2B303B')
                cell.alignment = left()
            elif c == 2:
                cell.font      = Font(name='Calibri', size=10,
                                      color='FFC62828' if is_req else 'FFE65100')
                cell.alignment = center()
            elif c == 3:
                cell.font      = Font(name='Calibri', bold=bold, size=10, color=color)
                cell.alignment = center()
            else:
                cell.font      = Font(name='Calibri', size=10, color='FF2B303B')
                cell.alignment = center()

    set_col_widths(ws, [28, 14, 16, 18, 14])
    ws.freeze_panes = 'A3'


def write_sheet_all_products(wb, df, attr_df):
    ws = wb.create_sheet('All Products')
    ws.sheet_view.showGridLines = False

    # Build column classification sets
    required_cols    = set(attr_df[attr_df['Type'] == 'Required']['Attribute'].tolist())
    recommended_cols = set(attr_df[attr_df['Type'] == 'Recommended']['Attribute'].tolist())

    cols = list(df.columns)

    # Header row
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font      = Font(name='Calibri', bold=True, size=10, color=C_WHITE)
        cell.fill      = fill(C_HEADER)
        cell.alignment = center(wrap=True)
        cell.border    = border

    # Pre-compute duplicate sets for content-quality checks
    dup_titles = set(df['title'].dropna()[df['title'].dropna().duplicated(keep=False)].tolist()) if 'title' in df.columns else set()
    dup_descs  = set(df['description'].dropna()[df['description'].dropna().duplicated(keep=False)].tolist()) if 'description' in df.columns else set()

    # Data rows with cell-level coloring
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(cols, 1):
            val      = row[col]
            is_blank = pd.isna(val) or str(val).strip() == ''
            cell     = ws.cell(row=r, column=c, value=(None if is_blank else val))
            cell.font      = Font(name='Calibri', size=9)
            cell.alignment = left()
            cell.border    = border

            if col in required_cols:
                if is_blank:
                    cell.fill = fill(C_RED_FILL)
                elif col in VALID_VALUES and str(val).lower() not in VALID_VALUES[col]:
                    cell.fill = fill(C_AMBER_FILL)
                elif col == 'google_product_category':
                    decoded = decode_category(str(val))
                    if str(val) in TAXONOMY and decoded.count('>') == 0:
                        cell.fill = fill(C_AMBER_FILL)
                    else:
                        cell.fill = fill(C_WHITE)
                elif col == 'title':
                    sval = str(val)
                    length = len(sval)
                    if length < 25 or length > 150 or sval in dup_titles:
                        cell.fill = fill(C_AMBER_FILL)
                    else:
                        cell.fill = fill(C_WHITE)
                elif col == 'description':
                    sval = str(val)
                    length = len(sval)
                    if length < 100 or length > 5000 or sval in dup_descs:
                        cell.fill = fill(C_AMBER_FILL)
                    else:
                        cell.fill = fill(C_WHITE)
                else:
                    cell.fill = fill(C_WHITE)

            elif col in recommended_cols:
                if is_blank:
                    cell.fill = fill(C_AMBER_FILL)
                elif col == 'product_type':
                    sval = str(val)
                    if '>' not in sval or sval.count('>') < 2:
                        cell.fill = fill(C_AMBER_FILL)
                    else:
                        cell.fill = fill(C_WHITE)
                else:
                    cell.fill = fill(C_WHITE)

            else:
                cell.fill = fill(C_GREY_FILL)

    # Column widths
    for c, col in enumerate(cols, 1):
        if col == 'id':
            ws.column_dimensions[get_column_letter(c)].width = 22
        elif col == 'title':
            ws.column_dimensions[get_column_letter(c)].width = 55
        elif col == 'description':
            ws.column_dimensions[get_column_letter(c)].width = 40
        else:
            ws.column_dimensions[get_column_letter(c)].width = 18

    ws.freeze_panes = 'A2'

    # Add a colour legend below the data
    legend_row = len(df) + 4
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=4)
    ws.cell(row=legend_row, column=1).value = 'COLOUR LEGEND'
    ws.cell(row=legend_row, column=1).font = Font(name='Calibri', bold=True, size=9, color='FF737B8C')

    legend_items = [
        (C_RED_FILL,   'Red — Required attribute missing or invalid'),
        (C_AMBER_FILL, 'Amber — Recommended attribute missing, or invalid value'),
        (C_GREY_FILL,  'Grey — Informational column (not audited)'),
        (C_WHITE,      'White — OK'),
    ]
    for i, (color, label) in enumerate(legend_items, legend_row + 1):
        cell = ws.cell(row=i, column=1, value='')
        cell.fill   = fill(color)
        cell.border = border
        lbl = ws.cell(row=i, column=2, value=label)
        lbl.font    = Font(name='Calibri', size=9, color='FF2B303B')


def write_sheet_product_issues(wb, product_issues_df):
    ws = wb.create_sheet('Product Issues')
    ws.sheet_view.showGridLines = False

    headers    = ['Product ID', 'Title', 'Attribute', 'Issue', 'Current Value']
    col_widths = [22, 45, 18, 42, 35]

    if product_issues_df.empty:
        ws.merge_cells('A1:E1')
        ws['A1'].value = 'No issues found — all products passed all checks.'
        ws['A1'].font  = Font(name='Calibri', size=11, color='FF2E7D32')
        ws['A1'].alignment = left()
        set_col_widths(ws, col_widths)
        return

    # ── Header row ─────────────────────────────────────────────────────────────
    write_header_row(ws, 1, headers, bg=C_HEADER)

    # ── Data rows ──────────────────────────────────────────────────────────────
    for r, (_, row) in enumerate(product_issues_df.iterrows(), 2):
        ws.row_dimensions[r].height = 18
        for c, key in enumerate(headers, 1):
            val  = row[key]
            cell = ws.cell(row=r, column=c, value=val)
            cell.border    = border
            cell.alignment = left(wrap=False)
            # Alternating row tint
            cell.fill = fill('FFF8F9FF' if r % 2 == 0 else 'FFFFFFFF')
            if c == 1:
                cell.font = Font(name='Calibri', size=9, bold=True, color='FF3664F4')
            elif c == 3:
                cell.font = Font(name='Calibri', size=9, bold=True, color='FF2B303B')
            else:
                cell.font = Font(name='Calibri', size=9, color='FF2B303B')

    # ── Wrap data in an Excel Table (native dropdown filters, no macros) ───────
    max_row = len(product_issues_df) + 1
    tbl_ref = f'A1:{get_column_letter(len(headers))}{max_row}'
    tbl = Table(displayName='ProductIssues', ref=tbl_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)

    set_col_widths(ws, col_widths)
    ws.freeze_panes = 'A2'


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='B6 Product Feed Audit — Phase 1')
    parser.add_argument('--feed',        required=True,       help='Path to TSV feed file')
    parser.add_argument('--brand',       required=True,       help='Brand or client name')
    parser.add_argument('--vertical',    default=None,        help='Override vertical detection')
    parser.add_argument('--output',      default=None,        help='Output directory (default: data/feeds/)')
    parser.add_argument('--detect-only', action='store_true', help='Print vertical detection and exit')
    args = parser.parse_args()

    print(f'\nB6 PRODUCT FEED AUDIT')
    print(f'Brand: {args.brand}')
    print(f'Feed:  {args.feed}\n')

    df = load_feed(args.feed)
    vertical = args.vertical if args.vertical else detect_vertical(df)
    vertical_label = VERTICAL_LABELS.get(vertical, vertical.title())

    print(f'Products: {len(df)}')
    print(f'Columns:  {len(df.columns)}')
    print(f'Vertical: {vertical_label}')

    if args.detect_only:
        print(f'\nColumn list: {df.columns.tolist()}')
        return

    print()

    skip_attrs = {attr for attr, check_fn in CONDITIONAL_ATTRS.items() if not check_fn(df)}
    attr_df   = audit_attributes(df, vertical, skip_attrs)
    summary   = build_summary(df, attr_df, args.brand, vertical)
    issues_df = build_issues(df, attr_df, vertical)

    wb = Workbook()
    wb.remove(wb.active)

    write_sheet_readme(wb, summary)
    write_sheet_summary(wb, summary, issues_df)
    write_sheet_attributes(wb, attr_df)
    write_sheet_all_products(wb, df, attr_df)
    product_issues_df = build_product_issues(df, attr_df)
    write_sheet_product_issues(wb, product_issues_df)

    output_dir = Path(args.output) if args.output else Path('data/feeds')
    output_dir.mkdir(parents=True, exist_ok=True)
    today      = datetime.today().strftime('%Y%m%d')
    brand_slug = args.brand.lower().replace(' ', '-')
    out_path   = output_dir / f'{today}-{brand_slug}-feed-audit.xlsx'

    wb.save(out_path)

    print(f'Issues found:         {len(issues_df)}')
    print(f'Products w/ issues:   {summary["products_with_issues"]}')
    print(f'Required attributes:  {summary["required_ok"]}/{summary["required_total"]} OK')
    print(f'Recommended present:  {summary["recommended_present"]}/{summary["recommended_total"]}')
    print(f'\nReport saved to: {out_path}')

    return str(out_path)


if __name__ == '__main__':
    main()
