#!/usr/bin/env python3
"""
B6 Search Term Profitability Audit
Beyond Six — Google Ads Agency

Classifies search terms by profitability tier and intent category using spend
thresholds and hybrid intent classification (regex + LLM) to avoid false positives.

Usage:
  python b6_search_term_audit.py --account=bsk --days=30 --target-cpa=50
  python b6_search_term_audit.py --account=bsk --days=90 --target-roas=160
  python b6_search_term_audit.py --account=bsk --start=2026-01-01 --end=2026-01-31 --target-cpa=auto
"""

import argparse
import io
import json
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta

# Fix Windows cp1252 encoding — force UTF-8 for terminal output
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
from pathlib import Path

import pandas as pd

# ─────────────────────────────────────────────────────────────────────────────
# Paths
# ─────────────────────────────────────────────────────────────────────────────

SKILL_DIR = Path(__file__).parent.parent
BRAIN_DIR = SKILL_DIR.parent.parent.parent  # scripts → skill → skills → .claude → brain
CONFIG_PATH = SKILL_DIR / "config.json"
ACCOUNTS_PATH = BRAIN_DIR / ".claude" / "accounts.json"
DATA_DIR = BRAIN_DIR / "data" / "google-ads"


# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

def load_config():
    with open(CONFIG_PATH) as f:
        return json.load(f)


def load_accounts():
    if not ACCOUNTS_PATH.exists():
        return {}
    with open(ACCOUNTS_PATH) as f:
        return json.load(f)


def resolve_account(alias, accounts):
    alias_lower = alias.lower()
    for key, account in accounts.items():
        if key.lower() == alias_lower:
            return key, account
        for a in account.get("aliases", []):
            if a.lower() == alias_lower:
                return key, account
    raise ValueError(f"Account '{alias}' not found in accounts.json")


# ─────────────────────────────────────────────────────────────────────────────
# Date helpers
# ─────────────────────────────────────────────────────────────────────────────

def get_date_range(days=None, start=None, end=None, timezone="UTC"):
    if start and end:
        return start, end
    if not days:
        days = 30
    today = datetime.now()
    end_date = (today - timedelta(days=1)).strftime("%Y-%m-%d")
    start_date = (today - timedelta(days=days)).strftime("%Y-%m-%d")
    return start_date, end_date


# ─────────────────────────────────────────────────────────────────────────────
# Google Ads API fetch
# ─────────────────────────────────────────────────────────────────────────────

def fetch_search_terms(customer_id, login_customer_id, start_date, end_date, min_impressions=0):
    """Pull search terms with performance metrics via Google Ads API."""
    try:
        from google.ads.googleads.client import GoogleAdsClient
        from google.ads.googleads.errors import GoogleAdsException
    except ImportError:
        print("ERROR: google-ads package not installed. Run: pip install google-ads")
        sys.exit(1)

    yaml_path = Path.home() / "google-ads.yaml"
    if not yaml_path.exists():
        print(f"ERROR: google-ads.yaml not found at {yaml_path}")
        sys.exit(1)

    client = GoogleAdsClient.load_from_storage(str(yaml_path))
    if login_customer_id and str(login_customer_id) != str(customer_id):
        client.login_customer_id = str(login_customer_id).replace("-", "")
    service = client.get_service("GoogleAdsService")

    cid = str(customer_id).replace("-", "")

    # Single query covering Search, Shopping, and PMax via campaign_search_term_view.
    # Using this resource instead of search_term_view avoids double-counting: the same
    # search term can appear from different campaigns (correct) but not from the same
    # campaign across two different API resources (which was the previous bug).
    query = f"""
        SELECT
            campaign_search_term_view.search_term,
            campaign.name,
            campaign.advertising_channel_type,
            metrics.impressions,
            metrics.clicks,
            metrics.cost_micros,
            metrics.conversions,
            metrics.conversions_value
        FROM campaign_search_term_view
        WHERE segments.date BETWEEN '{start_date}' AND '{end_date}'
          AND metrics.cost_micros > 0
        ORDER BY metrics.cost_micros DESC
    """

    rows = []
    try:
        for batch in service.search_stream(customer_id=cid, query=query):
            for row in batch.results:
                clicks = row.metrics.clicks
                impr   = row.metrics.impressions
                rows.append({
                    "search_term":  row.campaign_search_term_view.search_term,
                    "campaign":     row.campaign.name,
                    "channel_type": row.campaign.advertising_channel_type.name,
                    "ad_group":     "",
                    "impressions":  impr,
                    "clicks":       clicks,
                    "cost":         row.metrics.cost_micros / 1_000_000,
                    "conversions":  row.metrics.conversions,
                    "conv_value":   row.metrics.conversions_value,
                    "ctr":          (clicks / impr * 100) if impr > 0 else 0,
                })
    except GoogleAdsException as ex:
        print(f"Google Ads API error: {ex}")
        sys.exit(1)

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Structural waste detection
# ─────────────────────────────────────────────────────────────────────────────

def detect_structural_waste(term, waste_patterns):
    """
    Returns the waste category name if the term matches a structural waste
    pattern, otherwise returns None. Bypasses spend threshold entirely.
    """
    term_lower = term.lower()
    for category, patterns in waste_patterns.items():
        for pattern in patterns:
            if re.search(r'\b' + re.escape(pattern) + r'\b', term_lower):
                return category
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Intent classification — hybrid (regex → LLM → cache)
# ─────────────────────────────────────────────────────────────────────────────

# Maps raw LLM label → (intent_category, funnel_stage)
_INTENT_META = {
    "brand":          ("Brand",       None),
    "competitors":    ("Competitors", None),
    "product":        ("Product",     None),
    "generic_top":    ("Generic",     "Top Funnel"),
    "generic_mid":    ("Generic",     "Mid Funnel"),
    "generic_bottom": ("Generic",     "Bottom Funnel"),
}

# Maps raw label → cluster_key (displayed in report / Excel)
_CLUSTER_KEY_MAP = {
    "brand":          "Brand",
    "competitors":    "Competitors",
    "product":        "Product",
    "generic_top":    "Generic",
    "generic_mid":    "Generic",
    "generic_bottom": "Generic",
}

_VALID_LABELS = set(_INTENT_META.keys())


def _regex_classify(term, brand_terms, product_terms, competitor_terms, config):
    """
    Fast regex pre-filter. Returns a label string or None (send to LLM).
    Priority: brand > product > competitors > generic_bottom > generic_top.
    Also matches space/hyphen-stripped variants (e.g. "yourbrand" → "your brand").
    """
    t = term.lower()
    t_nospace = t.replace(" ", "").replace("-", "")

    for bt in brand_terms:
        bt_l = bt.lower()
        if re.search(r'\b' + re.escape(bt_l) + r'\b', t):
            return "brand"
        if bt_l.replace(" ", "").replace("-", "") in t_nospace:
            return "brand"

    for pt in product_terms:
        pt_l = pt.lower()
        if re.search(r'\b' + re.escape(pt_l) + r'\b', t):
            return "product"
        if pt_l.replace(" ", "").replace("-", "") in t_nospace:
            return "product"

    for ct in competitor_terms:
        ct_l = ct.lower()
        if re.search(r'\b' + re.escape(ct_l) + r'\b', t):
            return "competitors"
        if ct_l.replace(" ", "").replace("-", "") in t_nospace:
            return "competitors"

    # Top signals only — bottom funnel requires context the LLM handles better
    # e.g. "what can I buy to deworm myself" has "buy" but is clearly informational
    for sig in config["intent_classification"]["top_signals"]:
        if re.search(r'\b' + re.escape(sig.lower()) + r'\b', t):
            return "generic_top"

    return None


def _classify_batch_llm(batch, system_prompt, client, max_tokens=8192):
    """Send one batch of terms to Claude Haiku. Returns {term: label} dict."""
    try:
        import anthropic
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=max_tokens,
            system=system_prompt,
            messages=[{"role": "user", "content": json.dumps(batch)}],
        )
        text = resp.content[0].text.strip()
        start = text.find("{")
        end   = text.rfind("}") + 1
        if start >= 0 and end > start:
            raw = json.loads(text[start:end])
            return {t: (v if v in _VALID_LABELS else "generic_mid") for t, v in raw.items()}
    except Exception:
        pass
    return {t: "generic_mid" for t in batch}


def classify_intent_llm(terms, account_name, brand_terms, product_terms, competitor_terms):
    """
    Classify terms via Claude Haiku in batches of 200, parallelised (5 workers).
    Returns {term: label} dict.
    """
    try:
        import anthropic
    except ImportError:
        print("ERROR: anthropic package not installed. Run: pip install anthropic")
        sys.exit(1)

    client = anthropic.Anthropic()
    MAX_TOKENS = 8192   # Haiku max; 200-term batches need ~2,500–4,000 tokens
    BATCH_SIZE = 100    # Smaller batches → safer JSON, faster per-batch latency
    brand_str   = ", ".join(brand_terms)      if brand_terms      else "none"
    product_str = ", ".join(product_terms)    if product_terms    else "none"
    comp_str    = ", ".join(competitor_terms) if competitor_terms else "none"

    # Extract distinctive words from brand/product terms to use as brand-signal prefixes.
    # Skip very short or common English words that would over-match.
    _COMMON = {"the", "and", "for", "with", "from", "best", "care", "life", "plus",
               "max", "pro", "ultra", "super", "pure", "bio", "eco", "dog", "cat"}
    all_name_words = set()
    for t in (brand_terms or []) + (product_terms or []):
        for w in t.lower().split():
            if len(w) >= 5 and w not in _COMMON:
                all_name_words.add(w)
    brand_words_str = ", ".join(sorted(all_name_words)) if all_name_words else "none"

    system_prompt = (
        f"You classify Google Ads search terms for {account_name}.\n\n"
        "Assign each term to exactly one label:\n"
        f"- brand: Searches for {account_name} or people associated with the brand\n"
        "- product: Searches for one of this brand's specific product lines (even with typos, abbreviations, or concatenations)\n"
        "- competitors: Searches for competing brands or their products\n"
        "- generic_top: Pure informational — 'what is X', 'how does X work', 'benefits of X', 'what can I take to treat X'\n"
        "- generic_mid: Research/consideration OR specific need/problem — USE AS DEFAULT when unsure\n"
        "- generic_bottom: Clear purchase intent — the searcher wants to BUY something right now\n"
        "  USE for: 'buy X', 'where to buy X', 'where can I buy X', 'X near me', 'X online store', 'order X'\n"
        "  DO NOT USE for: 'what can I buy to treat X' (buy is instrumental, goal is the treatment → generic_top)\n"
        "  DO NOT USE for: 'should I buy X', 'is X worth buying' (still deciding → generic_mid)\n\n"
        f"Brand names: {brand_str}\n"
        f"Product lines: {product_str}\n"
        f"Competitor brands: {comp_str}\n\n"
        "Rules:\n"
        "- Typos, abbreviations, and concatenations of the listed names still count (e.g. 'acmebrnd' = 'acme brand' → brand)\n"
        f"- These words are distinctive to this brand/products: {brand_words_str}\n"
        "  If a search term uses one of these words as a modifier, lean toward brand or product even if the exact product isn't listed\n"
        "- Default to generic_mid when intent is unclear\n"
        "- Input is a JSON array of terms. Return ONLY valid JSON: {\"term\": \"label\", ...}"
    )

    batch_size = BATCH_SIZE
    batches    = [terms[i:i + batch_size] for i in range(0, len(terms), batch_size)]
    results    = {}

    with ThreadPoolExecutor(max_workers=5) as ex:
        futures = {ex.submit(_classify_batch_llm, b, system_prompt, client): b for b in batches}
        for fut in as_completed(futures):
            results.update(fut.result())

    return results


def classify_intent_cached(df, account_key, account_name, brand_terms, product_terms,
                           competitor_terms, config):
    """
    Hybrid classification pipeline:
      1. Regex pre-filter (free, instant) for brand/product/competitor/bottom/top signals
      2. LLM for unmatched terms with cost > 0
      3. Default generic_mid for zero-spend unmatched terms
      4. Disk cache — LLM results saved; subsequent runs reuse cache

    Returns {term: label} dict for every term in df.
    """
    cache_path = DATA_DIR / account_key / "intent-cache.json"
    cache = {}
    if cache_path.exists():
        try:
            with open(cache_path) as f:
                cache = json.load(f)
        except Exception:
            pass

    spend_map = dict(zip(df["search_term"], df["cost"]))
    results   = {}
    needs_llm = []

    for term in df["search_term"]:
        label = _regex_classify(term, brand_terms, product_terms, competitor_terms, config)
        if label:
            results[term] = label
        elif term in cache:
            results[term] = cache[term]
        elif spend_map.get(term, 0) > 0:
            needs_llm.append(term)
        else:
            results[term] = "generic_mid"

    if needs_llm:
        n_batches = (len(needs_llm) + 199) // 200
        print(f"  Classifying {len(df):,} terms...  "
              f"(LLM: {len(needs_llm):,} terms, {n_batches} batches)")
        llm_results = classify_intent_llm(needs_llm, account_name, brand_terms, product_terms, competitor_terms)
        results.update(llm_results)
        cache.update(llm_results)
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        with open(cache_path, "w") as f:
            json.dump(cache, f, indent=2)
    else:
        print(f"  Classifying {len(df):,} terms...  (LLM: 0, all pre-filtered or cached)")

    return results


def build_intent_clusters(df, account_key, account, config):
    """
    Classify every search term by intent and build cluster_key / cluster_size columns.
    Also adds intent_category and funnel_stage columns.
    Funnel stage is applied to ALL terms (brand, product, competitors, generic).
    """
    account_name     = account.get("name", account_key)
    brand_terms      = account.get("brand_terms") or [account_name]
    product_terms    = account.get("product_terms") or []
    competitor_terms = account.get("competitor_terms") or []

    label_map = classify_intent_cached(
        df, account_key, account_name, brand_terms, product_terms, competitor_terms, config
    )

    # Funnel stage from LLM label (generic_top/mid/bottom) or regex for non-generic
    _LABEL_FUNNEL = {
        "generic_top":    "Top Funnel",
        "generic_mid":    "Mid Funnel",
        "generic_bottom": "Bottom Funnel",
    }

    def _get_funnel(term):
        label = label_map.get(term, "generic_mid")
        if label in _LABEL_FUNNEL:
            return _LABEL_FUNNEL[label]
        # brand / product / competitors — check explicit signals first
        t = term.lower()
        for sig in config["intent_classification"]["top_signals"]:
            if re.search(r'\b' + re.escape(sig.lower()) + r'\b', t):
                return "Top Funnel"
        bottom = config["intent_classification"].get("bottom_signals", [])
        for sig in bottom:
            if re.search(r'\b' + re.escape(sig.lower()) + r'\b', t):
                return "Bottom Funnel"
        # Default by category: brand and product searches are purchase-intent by nature;
        # competitor searches are still in consideration phase
        if label in ("brand", "product"):
            return "Bottom Funnel"
        return "Mid Funnel"

    df = df.copy()
    df["cluster_key"]     = df["search_term"].map(
        lambda t: _CLUSTER_KEY_MAP.get(label_map.get(t, "generic_mid"), "Generic")
    )
    df["intent_category"] = df["search_term"].map(
        lambda t: _INTENT_META.get(label_map.get(t, "generic_mid"), ("Generic", "Mid Funnel"))[0]
    )
    df["funnel_stage"]    = df["search_term"].map(_get_funnel)

    cluster_sizes = df.groupby("cluster_key").size().rename("cluster_size")
    df = df.merge(cluster_sizes, on="cluster_key", how="left")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Profitability classification
# ─────────────────────────────────────────────────────────────────────────────

def classify_terms(df, spend_threshold, waste_patterns, strip_modifiers,
                   target_cpa=None, target_roas=None, high_cpa_multiplier=3.0):
    """
    Apply profitability tiers to every search term.

    Supports both CPA and ROAS targets. Pass one or the other.

    Order of classification:
    1. Structural waste  — pattern match, no spend threshold
    2. Profitable        — spend >= threshold AND above target metric
    3. Unprofitable      — spend >= threshold AND below target metric
    4. Cluster signal    — untested individually but cluster is profitable
    5. Untested          — everything else (insufficient data)
    """
    # Add computed columns
    df["cpa"] = df.apply(
        lambda r: r["cost"] / r["conversions"] if r["conversions"] > 0 else None, axis=1
    )
    df["roas"] = df.apply(
        lambda r: r["conv_value"] / r["cost"] if r["cost"] > 0 else None, axis=1
    )
    df["wasted_spend"] = 0.0
    df["tier"] = "untested"
    df["waste_subcategory"] = None

    # ── Step 1: Structural waste ──────────────────────────────────────────────
    # Only flag structural waste if the term actually spent money.
    # $0 terms with irrelevant patterns have nothing to action — leave as untested.
    for idx, row in df.iterrows():
        waste_cat = detect_structural_waste(row["search_term"], waste_patterns)
        if waste_cat and row["cost"] > 0:
            df.at[idx, "tier"] = "waste"
            df.at[idx, "waste_subcategory"] = waste_cat
            df.at[idx, "wasted_spend"] = row["cost"]

    remaining = df[df["tier"] == "untested"]

    if target_roas is not None:
        # ── ROAS mode ─────────────────────────────────────────────────────────
        profitable_mask = (
            (remaining["cost"] >= spend_threshold) &
            (remaining["conversions"] > 0) &
            (remaining["roas"] >= target_roas)
        )
        df.loc[remaining[profitable_mask].index, "tier"] = "profitable"

        remaining = df[df["tier"] == "untested"]
        unprofitable_mask = (
            (remaining["cost"] >= spend_threshold) &
            (
                (remaining["conversions"] == 0) |
                (remaining["roas"] < target_roas)
            )
        )
    else:
        # ── CPA mode ──────────────────────────────────────────────────────────
        high_cpa_threshold = target_cpa * high_cpa_multiplier
        profitable_mask = (
            (remaining["cost"] >= spend_threshold) &
            (remaining["conversions"] > 0) &
            (remaining["cpa"] <= target_cpa)
        )
        df.loc[remaining[profitable_mask].index, "tier"] = "profitable"

        remaining = df[df["tier"] == "untested"]
        unprofitable_mask = (
            (remaining["cost"] >= spend_threshold) &
            (
                (remaining["conversions"] == 0) |
                (remaining["cpa"] > high_cpa_threshold)
            )
        )

    # ── Step 3: Mark unprofitable ─────────────────────────────────────────────
    df.loc[remaining[unprofitable_mask].index, "tier"] = "unprofitable"
    df.loc[remaining[unprofitable_mask].index, "wasted_spend"] = remaining[unprofitable_mask]["cost"]

    return df


# ─────────────────────────────────────────────────────────────────────────────
# CPA auto-calculation
# ─────────────────────────────────────────────────────────────────────────────

def calculate_account_cpa(df):
    """Derive target CPA from total account spend / total conversions."""
    total_spend = df["cost"].sum()
    total_conv = df["conversions"].sum()
    if total_conv == 0:
        print("WARNING: No conversions in this period. Cannot calculate CPA automatically.")
        print("         Please provide --target-cpa manually.")
        sys.exit(1)
    return total_spend / total_conv


def calculate_avg_order_value(df):
    """Derive average order value from total conv_value / total conversions."""
    total_conv = df["conversions"].sum()
    total_value = df["conv_value"].sum()
    if total_conv == 0 or total_value == 0:
        return None
    return total_value / total_conv


# ─────────────────────────────────────────────────────────────────────────────
# Terminal report
# ─────────────────────────────────────────────────────────────────────────────

def print_report(df, account_name, start_date, end_date, target_cpa, currency, config, spend_threshold=None, target_roas=None):
    width = 68
    top_n = config["output"]["top_waste_count"]
    top_clusters = config["output"]["top_clusters_count"]
    cluster_min = config["output"]["cluster_min_terms"]
    if spend_threshold is None:
        spend_threshold = target_cpa * config["thresholds"]["spend_multiplier"] if target_cpa else 0

    def fmt_currency(amount):
        return f"{currency}{amount:,.2f}"

    print()
    print("═" * width)
    print("  B6 SEARCH TERM PROFITABILITY AUDIT")
    print(f"  {account_name} | {start_date} – {end_date}")
    print("═" * width)
    print()
    if target_roas:
        print(f"  TARGET ROAS: {target_roas*100:.0f}%  |  SPEND THRESHOLD: {fmt_currency(spend_threshold)}")
    else:
        print(f"  TARGET CPA: {fmt_currency(target_cpa)}  |  SPEND THRESHOLD: {fmt_currency(spend_threshold)}")
    print()

    # ── Summary table ─────────────────────────────────────────────────────────
    tier_order = ["profitable", "unprofitable", "untested", "waste"]
    tier_labels = {
        "profitable":  "Profitable  ",
        "unprofitable":"Unprofitable",
        "untested":    "Untested    ",
        "waste":       "Waste       ",
    }

    print("  PROFITABILITY SUMMARY")
    print(f"  {'Tier':<20} {'Terms':>6}  {'Spend':>10}  {'Conv':>6}  {'Notes'}")
    print("  " + "─" * (width - 2))

    for tier in tier_order:
        subset = df[df["tier"] == tier]
        if subset.empty:
            continue
        terms = len(subset)
        spend = subset["cost"].sum()
        conv = subset["conversions"].sum()
        wasted = subset["wasted_spend"].sum()

        notes = ""
        if tier == "profitable":
            avg_cpa = spend / conv if conv > 0 else 0
            notes = f"avg CPA {fmt_currency(avg_cpa)}"
        elif tier in ("unprofitable", "waste"):
            notes = f"{fmt_currency(wasted)} wasted"

        print(f"  {tier_labels[tier]} {terms:>6}  {fmt_currency(spend):>10}  {conv:>6.0f}  {notes}")

    total_wasted = df["wasted_spend"].sum()
    print()
    print(f"  TOTAL WASTED SPEND: {fmt_currency(total_wasted)}")
    print()

    # ── Intent breakdown ──────────────────────────────────────────────────────
    if "cluster_key" in df.columns and df["cluster_key"].notna().any():
        print("─" * width)
        print("  INTENT BREAKDOWN")
        print("─" * width)
        print(f"  {'Category':<20} {'Terms':>6}  {'Spend':>10}  {'Conv':>6}  {'ROAS':>6}")
        print("  " + "─" * (width - 2))

        generic_keys = {"Generic"}
        for label, keys in [
            ("Brand",       {"Brand"}),
            ("Competitors", {"Competitors"}),
            ("Product",     {"Product"}),
            ("Generic",     generic_keys),
        ]:
            sub = df[df["cluster_key"].isin(keys)]
            if sub.empty:
                continue
            spend = sub["cost"].sum()
            conv  = sub["conversions"].sum()
            val   = sub["conv_value"].sum()
            roas  = val / spend if spend > 0 and val > 0 else None
            print(f"  {label:<20} {len(sub):>6}  {fmt_currency(spend):>10}  {conv:>6.0f}  {f'{roas:.1f}x' if roas else '—':>6}")

        print()
        print("  FUNNEL BREAKDOWN  (all terms)")
        print("─" * width)
        print(f"  {'Funnel Stage':<20} {'Terms':>6}  {'Spend':>10}  {'Conv':>6}  {'ROAS':>6}")
        print("  " + "─" * (width - 2))

        for label in ["Top Funnel", "Mid Funnel", "Bottom Funnel"]:
            sub = df[df["funnel_stage"] == label]
            if sub.empty:
                continue
            spend = sub["cost"].sum()
            conv  = sub["conversions"].sum()
            val   = sub["conv_value"].sum()
            roas  = val / spend if spend > 0 and val > 0 else None
            print(f"  {label:<20} {len(sub):>6}  {fmt_currency(spend):>10}  {conv:>6.0f}  {f'{roas:.1f}x' if roas else '—':>6}")

        print()

    print("═" * width)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

# Tier colour palette
_TIER_FILL = {
    "profitable":  "C6EFCE",
    "unprofitable":"FFC7CE",
    "untested":    "FFEB9C",
    "waste":       "FCE4D6",
}
_TIER_FONT = {
    "profitable":  "375623",
    "unprofitable":"9C0006",
    "untested":    "7D6608",
    "waste":       "843C0C",
}

def generate_recommendations(df_agg, target_cpa, target_roas, spend_threshold, currency):
    """Analyse df_agg and return a list of actionable recommendations.

    Each item: { priority, category, finding, action }
    Priority: "High" | "Medium" | "Low"
    """
    recs = []

    def fc(amount):
        return f"{currency}{amount:,.2f}"

    def roas_str(val):
        return f"{val:.1f}x" if val else "0.0x"

    total_spend = df_agg["cost"].sum()
    total_conv  = df_agg["conversions"].sum()

    # ── Waste terms ───────────────────────────────────────────────────────────
    waste = df_agg[df_agg["tier"] == "waste"]
    waste_spend = waste["cost"].sum()
    if len(waste) > 0 and waste_spend > 0:
        recs.append({
            "priority": "High",
            "category": "Negative Keywords",
            "finding":  f"{len(waste)} waste terms spent {fc(waste_spend)} with 0 conversions. These match irrelevant queries — job searches, Reddit, DIY, etc.",
            "action":   "Add all waste terms as negatives immediately. They have no conversion potential regardless of bid.",
        })

    # ── Unprofitable terms — per term, split by intent ────────────────────────
    unprof = df_agg[df_agg["tier"] == "unprofitable"].sort_values("cost", ascending=False)
    for _, row in unprof.iterrows():
        term  = row["search_term"]
        spend = row["cost"]
        ck    = row.get("cluster_key", "Generic") or "Generic"
        if ck == "Product":
            recs.append({
                "priority": "High",
                "category": "Landing Page / Offer",
                "finding":  f'"{term}" spent {fc(spend)} above threshold with 0 conversions. The intent is product-level — the query is right.',
                "action":   "This points to a landing page or offer issue, not the term itself. Review the landing page for relevance, offer strength, and friction.",
            })
        elif ck == "Competitors":
            recs.append({
                "priority": "Medium",
                "category": "Negative Keywords",
                "finding":  f'Competitor term "{term}" spent {fc(spend)} above threshold with 0 conversions.',
                "action":   "Assess whether this competitor is worth bidding on. If keeping, test a dedicated competitor landing page with direct comparison messaging.",
            })
        else:
            recs.append({
                "priority": "Medium",
                "category": "Negative Keywords",
                "finding":  f'"{term}" spent {fc(spend)} above threshold with 0 conversions. Generic intent with no return.',
                "action":   "Add as negative or reduce bids significantly. The query isn't earning its place.",
            })

    # ── Funnel analysis ───────────────────────────────────────────────────────
    funnel = {}
    for stage in ["Top Funnel", "Mid Funnel", "Bottom Funnel"]:
        sub = df_agg[df_agg["funnel_stage"] == stage] if "funnel_stage" in df_agg.columns else None
        if sub is not None and not sub.empty:
            sp  = sub["cost"].sum()
            val = sub["conv_value"].sum()
            funnel[stage] = {
                "spend": sp,
                "roas":  val / sp if sp > 0 else 0,
                "conv":  sub["conversions"].sum(),
                "terms": len(sub),
            }

    top = funnel.get("Top Funnel")
    mid = funnel.get("Mid Funnel")
    bot = funnel.get("Bottom Funnel")

    if top and mid and top["roas"] > mid["roas"] and mid["spend"] > 0:
        recs.append({
            "priority": "Medium",
            "category": "Funnel Strategy",
            "finding":  f"Top funnel converts at {roas_str(top['roas'])} ROAS vs {roas_str(mid['roas'])} for mid funnel. Top funnel outperforming mid is unusual.",
            "action":   "Create dedicated landing pages for top funnel keywords and test increased budget there. Mid funnel terms — which should be closer to purchase — need attention: check landing page relevance and offer strength.",
        })
    elif mid and bot and mid["roas"] < bot["roas"] * 0.5 and bot["spend"] > 0:
        recs.append({
            "priority": "Medium",
            "category": "Funnel Strategy",
            "finding":  f"Mid funnel ROAS ({roas_str(mid['roas'])}) is well below bottom funnel ({roas_str(bot['roas'])}). Consideration-stage traffic is underperforming.",
            "action":   "Review mid funnel ad-to-page relevance. Searchers in the comparison phase need differentiation and social proof — not generic product pages.",
        })

    if bot and bot["roas"] > 0:
        eff_target = (target_roas * 2) if target_roas else ((target_cpa * 2) if target_cpa else None)
        bot_is_strong = eff_target is None or (
            target_roas and bot["roas"] > target_roas * 2
        ) or (
            target_cpa and bot["conv"] > 0 and bot["spend"] / bot["conv"] < target_cpa * 0.5
        )
        if bot_is_strong:
            recs.append({
                "priority": "Low",
                "category": "Budget Allocation",
                "finding":  f"Bottom funnel terms are your strongest performers at {roas_str(bot['roas'])} ROAS across {bot['terms']} terms.",
                "action":   "Ensure bottom funnel campaigns are uncapped. Any budget ceiling here is limiting your highest-intent traffic.",
            })

    if bot and total_spend > 0 and bot["spend"] / total_spend > 0.5:
        bot_pct = bot["spend"] / total_spend * 100
        recs.append({
            "priority": "Medium",
            "category": "Incrementality Risk",
            "finding":  f"Bottom funnel terms account for {bot_pct:.0f}% of total spend. Heavy concentration here often means over-investing in branded and product keywords.",
            "action":   "Run an incrementality test on brand and bottom-funnel terms. Users searching your brand name may convert regardless of whether an ad is shown — meaning some of this spend may not be driving incremental revenue.",
        })

    # ── Intent mix ────────────────────────────────────────────────────────────
    intent = {}
    if "cluster_key" in df_agg.columns:
        for ck in ["Brand", "Competitors", "Product", "Generic"]:
            sub = df_agg[df_agg["cluster_key"] == ck]
            if not sub.empty:
                sp  = sub["cost"].sum()
                val = sub["conv_value"].sum()
                intent[ck] = {
                    "spend": sp,
                    "roas":  val / sp if sp > 0 else 0,
                    "conv":  sub["conversions"].sum(),
                    "terms": len(sub),
                }

    comp = intent.get("Competitors")
    if comp and comp["roas"] < 1.0 and comp["spend"] > 50:
        recs.append({
            "priority": "Medium",
            "category": "Intent Mix",
            "finding":  f"Competitor terms spent {fc(comp['spend'])} at {roas_str(comp['roas'])} ROAS — below break-even.",
            "action":   "Pause or restructure competitor campaigns. If keeping, build a dedicated competitor page with direct comparison and a strong reason to switch.",
        })

    generic = intent.get("Generic")
    if generic and total_spend > 0 and generic["spend"] / total_spend > 0.5 and generic["roas"] < 1.5:
        recs.append({
            "priority": "Medium",
            "category": "Intent Mix",
            "finding":  f"Generic terms represent {generic['spend']/total_spend*100:.0f}% of spend at {roas_str(generic['roas'])} ROAS. Most are broad, low-intent queries.",
            "action":   "Shift budget toward Product and Brand terms, which show stronger intent. Tighten match types on generic terms or apply bid adjustments.",
        })

    brand = intent.get("Brand")
    if brand and brand["conv"] > 0 and total_conv > 0:
        brand_conv_pct = brand["conv"] / total_conv * 100
        if brand_conv_pct > 40:
            recs.append({
                "priority": "Low",
                "category": "Budget Allocation",
                "finding":  f"Brand terms drive {brand_conv_pct:.0f}% of all conversions at {roas_str(brand['roas'])} ROAS.",
                "action":   "Ensure brand campaigns are fully uncapped. Losing impression share here means ceding high-intent, low-cost conversions to competitors.",
            })

    # ── Data coverage ─────────────────────────────────────────────────────────
    untested = df_agg[df_agg["tier"] == "untested"]
    unt_spend = untested["cost"].sum()
    unt_pct = unt_spend / total_spend * 100 if total_spend > 0 else 0
    if unt_pct > 60:
        recs.append({
            "priority": "Low",
            "category": "Data Coverage",
            "finding":  f"{unt_pct:.0f}% of spend ({fc(unt_spend)}) is in untested terms — each below the {fc(spend_threshold)} threshold.",
            "action":   "Most terms haven't spent enough to be judged. Use --min-spend to lower the threshold, or extend the date range to give more terms enough data.",
        })

    return recs


def export_excel(df, account_name, account_key, start_date, end_date,
                 target_cpa, target_roas, spend_threshold, currency, config, output_dir):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")

    # ── Style helpers ─────────────────────────────────────────────────────────
    NAVY   = "141D2E"
    COBALT = "3664F4"
    WHITE  = "FFFFFF"
    DGRAY  = "2B303B"
    LGRAY  = "F9F9F9"

    def tfill(tier):
        return PatternFill("solid", fgColor=_TIER_FILL.get(tier, "FFFFFF"))

    def tfont(tier, bold=False, size=9):
        return Font(color=_TIER_FONT.get(tier, "000000"), bold=bold, size=size)

    def hfill(color):
        return PatternFill("solid", fgColor=color)

    def hfont(color=WHITE, bold=True, size=9):
        return Font(color=color, bold=bold, size=size)

    def center(indent=0):
        return Alignment(horizontal="center", vertical="center", indent=indent)

    def left(indent=0):
        return Alignment(horizontal="left", vertical="center", indent=indent)

    def safe(val):
        """Return None instead of NaN so openpyxl doesn't choke."""
        if val is None:
            return None
        try:
            if pd.isna(val):
                return None
        except (TypeError, ValueError):
            pass
        return val

    def fmt_c(amount):
        return f"{currency}{amount:,.2f}" if amount else f"{currency}0.00"

    def write_header_row(ws, row_num, labels, bg=COBALT, fg=WHITE, height=18):
        for col, label in enumerate(labels, 1):
            c = ws.cell(row=row_num, column=col, value=label)
            c.fill = hfill(bg)
            c.font  = hfont(fg, bold=True, size=9)
            c.alignment = center()
        ws.row_dimensions[row_num].height = height

    def set_widths(ws, widths):
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

    tier_labels = {
        "profitable":  "Profitable",
        "unprofitable":"Unprofitable",
        "untested":    "Untested",
        "waste":       "Waste",
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    NCOLS = 6  # columns A–F used throughout Summary

    def section_hdr(ws, r, title):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        c = ws.cell(r, 1, title)
        c.fill = hfill("E8F2FF"); c.font = Font(bold=True, size=10, color=DGRAY)
        c.alignment = left(1)
        ws.row_dimensions[r].height = 20
        return r + 1

    def spacer(ws, r, height=10):
        ws.row_dimensions[r].height = height
        return r + 1

    # ── Build aggregated dataframe (one row per unique search term) ────────────
    def _dominant(grp, col):
        """Return value of col from the row with the highest cost in the group."""
        return grp.loc[grp["cost"].idxmax(), col]

    agg_rows = []
    for term, grp in df.groupby("search_term", sort=False):
        agg_rows.append({
            "search_term":     term,
            "impressions":     grp["impressions"].sum(),
            "clicks":          grp["clicks"].sum(),
            "cost":            grp["cost"].sum(),
            "conversions":     grp["conversions"].sum(),
            "conv_value":      grp["conv_value"].sum(),
            "wasted_spend":    grp["wasted_spend"].sum(),
            "tier":            _dominant(grp, "tier"),
            "intent_category": _dominant(grp, "intent_category") if "intent_category" in grp.columns else None,
            "funnel_stage":    _dominant(grp, "funnel_stage")    if "funnel_stage"    in grp.columns else None,
            "cluster_key":     _dominant(grp, "cluster_key")     if "cluster_key"     in grp.columns else None,
            "channels":        ", ".join(sorted(grp["channel_type"].unique())),
            "campaigns":       f"{grp['campaign'].nunique()} campaign{'s' if grp['campaign'].nunique() > 1 else ''}",
        })
    df_agg = pd.DataFrame(agg_rows)
    df_agg["ctr"]  = df_agg.apply(
        lambda r: r["clicks"] / r["impressions"] * 100 if r["impressions"] > 0 else 0, axis=1)
    df_agg["cpc"]  = df_agg.apply(
        lambda r: r["cost"] / r["clicks"] if r["clicks"] > 0 else None, axis=1)
    df_agg["cpa"]  = df_agg.apply(
        lambda r: r["cost"] / r["conversions"] if r["conversions"] > 0 else None, axis=1)
    df_agg["roas"] = df_agg.apply(
        lambda r: r["conv_value"] / r["cost"] if r["cost"] > 0 else None, axis=1)
    df_agg = df_agg.sort_values("cost", ascending=False).reset_index(drop=True)

    recs = generate_recommendations(df_agg, target_cpa, target_roas, spend_threshold, currency)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    r = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(r, 1, "B6 SEARCH TERM PROFITABILITY AUDIT")
    c.fill = hfill(NAVY); c.font = Font(bold=True, size=14, color=WHITE)
    c.alignment = center()
    ws.row_dimensions[r].height = 32
    r += 1

    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(r, 1, f"{account_name}  |  {start_date} – {end_date}")
    c.fill = hfill(NAVY); c.font = Font(size=10, color="AAAAAA")
    c.alignment = center()
    ws.row_dimensions[r].height = 20
    r += 1
    r = spacer(ws, r, 10)

    # ── KPI scorecards ────────────────────────────────────────────────────────
    total_spend  = df_agg["cost"].sum()
    total_value  = df_agg["conv_value"].sum()
    total_conv   = df_agg["conversions"].sum()
    total_terms  = len(df_agg)
    total_wasted = df_agg["wasted_spend"].sum()
    waste_pct    = total_wasted / total_spend * 100 if total_spend > 0 else 0
    overall_roas = total_value / total_spend if total_spend > 0 and total_value > 0 else None
    prof_count   = int((df_agg["tier"] == "profitable").sum())

    # Waste tile urgency color (traffic-light)
    if waste_pct > 10:
        w_bg, w_fg = "9B1C1C", WHITE   # deep red  — danger
    elif waste_pct > 5:
        w_bg, w_fg = "C0392B", WHITE   # red       — high
    elif waste_pct > 2:
        w_bg, w_fg = "D97706", WHITE   # amber     — moderate
    else:
        w_bg, w_fg = "1A5C38", WHITE   # green     — healthy

    # KPI row — aligns with table column headers: [Label | Terms | Spend | Revenue | ROAS | Conversions]
    # (label, value, label_bg, label_fg, value_bg, value_fg)
    kpi_tiles = [
        ("TOTAL",        f"{account_name}",                                NAVY,     WHITE,   "E8F2FF", NAVY),
        ("Terms",        f"{total_terms:,}",                               DGRAY,    WHITE,   "F2F2F2", DGRAY),
        ("Spend",        fmt_c(total_spend),                               COBALT,   WHITE,   "EEF3FF", "141D2E"),
        ("Revenue",      fmt_c(total_value) if total_value > 0 else "—",   "1A5C38", WHITE,   "E2EFDA", "1A5C38"),
        ("ROAS",         f"{overall_roas:.1f}x" if overall_roas else "—",  "17375E", WHITE,   "D9E1F2", "17375E"),
        ("Conversions",  f"{int(total_conv):,}",                           "404040", WHITE,   "F9F9F9", "404040"),
    ]

    # Label row
    for ci, tile in enumerate(kpi_tiles, 1):
        label, _, lbg, lfg, _, _ = tile
        c = ws.cell(r, ci, label)
        c.fill = hfill(lbg); c.font = Font(color=lfg, bold=True, size=8)
        c.alignment = center()
    # Waste section header at I4
    c = ws.cell(r, 9, "WASTE HEALTH")
    c.fill = hfill(w_bg); c.font = Font(color=w_fg, bold=True, size=8)
    c.alignment = center()
    ws.row_dimensions[r].height = 16
    r += 1

    # Value row
    for ci, tile in enumerate(kpi_tiles, 1):
        _, val, _, _, vbg, vfg = tile
        c = ws.cell(r, ci, val)
        c.fill = hfill(vbg); c.font = Font(color=vfg, bold=True, size=13)
        c.alignment = center()
    # Tile 1 at I5 — Wasted Ad Spend (label + value in one cell)
    c = ws.cell(r, 9, f"WASTED AD SPEND\n{fmt_c(total_wasted)}")
    c.fill = hfill(w_bg); c.font = Font(color=w_fg, bold=True, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 40
    r += 1

    # Tile 2 at I6 — Waste % (label + value in one cell, replaces spacer)
    c = ws.cell(r, 9, f"WASTE %\n{waste_pct:.1f}%")
    c.fill = hfill(w_bg); c.font = Font(color=w_fg, bold=True, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 40
    r += 1

    # ── Profitability summary ─────────────────────────────────────────────────
    r = section_hdr(ws, r, "PROFITABILITY SUMMARY")
    write_header_row(ws, r, ["Tier", "Terms", "Spend", "Revenue", "ROAS", "Conversions"])
    r += 1

    tier_order = ["profitable", "unprofitable", "untested", "waste"]

    for tier in tier_order:
        sub = df_agg[df_agg["tier"] == tier]
        if sub.empty:
            continue
        spend  = sub["cost"].sum()
        conv   = sub["conversions"].sum()
        val    = sub["conv_value"].sum()
        roas   = val / spend if spend > 0 and val > 0 else None
        wasted = sub["wasted_spend"].sum()
        notes  = ""
        if tier == "profitable":
            avg_cpa = spend / conv if conv > 0 else 0
            notes = f"avg CPA {fmt_c(avg_cpa)}"
        elif tier in ("unprofitable", "waste"):
            notes = f"{fmt_c(wasted)} wasted"

        vals = [tier_labels[tier], len(sub), fmt_c(spend),
                fmt_c(val) if val > 0 else "—",
                f"{roas:.1f}x" if roas else "—",
                f"{conv:.0f}"]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            c.fill = tfill(tier); c.font = tfont(tier)
            c.alignment = left(1) if col == 1 else center()
        r += 1

    # ── Intent breakdown ──────────────────────────────────────────────────────
    if "cluster_key" in df_agg.columns and df_agg["cluster_key"].notna().any():
        r = spacer(ws, r, 16)
        r = section_hdr(ws, r, "INTENT BREAKDOWN")
        write_header_row(ws, r, ["Category", "Terms", "Spend", "Revenue", "ROAS", "Conversions"])
        r += 1

        _INTENT_GROUP_BG = {
            "Brand":       "D9E1F2",
            "Competitors": "FCE4D6",
            "Product":     "E2EFDA",
            "Generic":     "FFEB9C",
        }
        _INTENT_GROUP_FG = {
            "Brand":       "1F3864",
            "Competitors": "843C0C",
            "Product":     "375623",
            "Generic":     "7D6608",
        }
        generic_keys = {"Generic"}
        for label, keys in [
            ("Brand",       {"Brand"}),
            ("Competitors", {"Competitors"}),
            ("Product",     {"Product"}),
            ("Generic",     generic_keys),
        ]:
            sub = df_agg[df_agg["cluster_key"].isin(keys)]
            if sub.empty:
                continue
            spend = sub["cost"].sum()
            conv  = sub["conversions"].sum()
            val   = sub["conv_value"].sum()
            roas  = val / spend if spend > 0 and val > 0 else None
            bg = _INTENT_GROUP_BG.get(label, "FFFFFF")
            fg = _INTENT_GROUP_FG.get(label, DGRAY)
            vals = [label, len(sub), fmt_c(spend),
                    fmt_c(val) if val > 0 else "—",
                    f"{roas:.1f}x" if roas else "—",
                    f"{conv:.0f}"]
            for col, v in enumerate(vals, 1):
                c = ws.cell(r, col, v)
                c.fill = hfill(bg); c.font = Font(color=fg, size=9)
                c.alignment = left(1) if col == 1 else center()
            r += 1

        r = spacer(ws, r, 16)

        # ── Funnel breakdown ──────────────────────────────────────────────────
        r = section_hdr(ws, r, "FUNNEL BREAKDOWN")
        write_header_row(ws, r, ["Funnel Stage", "Terms", "Spend", "Revenue", "ROAS", "Conversions"])
        r += 1

        _FUNNEL_SUM_BG = {"Top Funnel": "EDEDED", "Mid Funnel": "FFEB9C", "Bottom Funnel": "C6EFCE"}
        _FUNNEL_SUM_FG = {"Top Funnel": "595959", "Mid Funnel": "7D6608", "Bottom Funnel": "375623"}
        for stage in ["Top Funnel", "Mid Funnel", "Bottom Funnel"]:
            sub = df_agg[df_agg["funnel_stage"] == stage]
            if sub.empty:
                continue
            spend = sub["cost"].sum()
            conv  = sub["conversions"].sum()
            val   = sub["conv_value"].sum()
            roas  = val / spend if spend > 0 and val > 0 else None
            bg = _FUNNEL_SUM_BG.get(stage, "FFFFFF")
            fg = _FUNNEL_SUM_FG.get(stage, DGRAY)
            vals = [stage, len(sub), fmt_c(spend),
                    fmt_c(val) if val > 0 else "—",
                    f"{roas:.1f}x" if roas else "—",
                    f"{conv:.0f}"]
            for col, v in enumerate(vals, 1):
                c = ws.cell(r, col, v)
                c.fill = hfill(bg); c.font = Font(color=fg, size=9)
                c.alignment = left(1) if col == 1 else center()
            r += 1

    set_widths(ws, {"A": 22, "B": 10, "C": 14, "D": 14, "E": 10, "F": 14, "G": 2, "H": 2, "I": 20})

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Profitability Summary
    # ══════════════════════════════════════════════════════════════════════════
    wsp = wb.create_sheet("Profitability Summary")
    wsp.sheet_view.showGridLines = False
    pr = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    wsp.merge_cells(f"A{pr}:L{pr}")
    c = wsp.cell(pr, 1, "PROFITABILITY SUMMARY")
    c.fill = hfill(NAVY); c.font = Font(bold=True, size=13, color=WHITE)
    c.alignment = center()
    wsp.row_dimensions[pr].height = 28
    pr += 1

    wsp.merge_cells(f"A{pr}:K{pr}")
    c = wsp.cell(pr, 1, f"{account_name}  |  {start_date} – {end_date}")
    c.fill = hfill(NAVY); c.font = Font(size=9, color="AAAAAA")
    c.alignment = center()
    wsp.row_dimensions[pr].height = 16
    pr += 1
    wsp.row_dimensions[pr].height = 8
    pr += 1

    # ── Tier summary table ────────────────────────────────────────────────────
    ps_hdr = ["Tier", "Terms", "Spend", "Clicks", "CTR", "CPC", "Conversions",
              "Conv Rate", "CPA", "Conv Value", "ROAS"]
    write_header_row(wsp, pr, ps_hdr, bg=COBALT, height=18)
    pr += 1

    for tier in ["profitable", "unprofitable", "untested", "waste"]:
        sub = df_agg[df_agg["tier"] == tier]
        if sub.empty:
            continue
        spend  = sub["cost"].sum()
        clicks = int(sub["clicks"].sum())
        impr   = sub["impressions"].sum()
        conv   = sub["conversions"].sum()
        val    = sub["conv_value"].sum()
        wasted = sub["wasted_spend"].sum()
        ctr    = sub["clicks"].sum() / impr * 100 if impr > 0 else 0
        cpc    = spend / clicks if clicks > 0 else None
        cpa    = spend / conv if conv > 0 else None
        conv_rate = conv / clicks * 100 if clicks > 0 else None
        roas   = val / spend if spend > 0 and val > 0 else None

        vals = [
            tier_labels[tier],
            len(sub),
            fmt_c(spend),
            clicks,
            f"{ctr:.2f}%",
            fmt_c(cpc) if cpc else "—",
            f"{conv:.0f}",
            f"{conv_rate:.2f}%" if conv_rate else "—",
            fmt_c(cpa) if cpa else "—",
            fmt_c(val) if val > 0 else "—",
            f"{roas:.1f}x" if roas else "—",
        ]
        for col, v in enumerate(vals, 1):
            c = wsp.cell(pr, col, v)
            c.fill = tfill(tier); c.font = tfont(tier, size=9)
            c.alignment = left(1) if col == 1 else center()
        pr += 1

    pr += 1  # spacer

    # ── Per-tier term detail ───────────────────────────────────────────────────
    detail_hdr = ["Search Term", "Spend", "Clicks", "CTR", "CPC",
                  "Conversions", "Conv Rate", "CPA", "Conv Value", "ROAS"]
    tier_section_bg = {
        "profitable":   "E2EFDA",
        "unprofitable": "FCE4D6",
        "untested":     "F2F2F2",
        "waste":        "FCE4D6",
    }
    tier_section_fg = {
        "profitable":   "375623",
        "unprofitable": "843C0C",
        "untested":     "595959",
        "waste":        "843C0C",
    }
    tier_section_label = {
        "profitable":   "PROFITABLE TERMS",
        "unprofitable": "UNPROFITABLE TERMS",
        "untested":     "UNTESTED TERMS",
        "waste":        "STRUCTURAL WASTE",
    }

    for tier in ["profitable", "unprofitable", "waste", "untested"]:
        sub = df_agg[df_agg["tier"] == tier].sort_values("cost", ascending=False)
        if sub.empty:
            continue

        # Section header
        wsp.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=len(detail_hdr))
        c = wsp.cell(pr, 1, f"{tier_section_label[tier]}  ({len(sub)} terms)")
        c.fill = hfill(tier_section_bg[tier])
        c.font = Font(bold=True, size=10, color=tier_section_fg[tier])
        c.alignment = left(1)
        wsp.row_dimensions[pr].height = 20
        pr += 1

        # Column headers
        write_header_row(wsp, pr, detail_hdr, bg=DGRAY, height=16)
        pr += 1

        for ri2, (_, row) in enumerate(sub.iterrows()):
            stripe = hfill(LGRAY) if ri2 % 2 == 0 else hfill("FFFFFF")
            spend2   = safe(row.get("cost"))
            clicks2  = int(float(safe(row.get("clicks")) or 0))
            impr2    = float(safe(row.get("impressions")) or 0)
            conv2    = float(safe(row.get("conversions")) or 0)
            val2     = float(safe(row.get("conv_value")) or 0)
            ctr2     = clicks2 / impr2 * 100 if impr2 > 0 else 0
            cpc2     = spend2 / clicks2 if clicks2 > 0 and spend2 else None
            cpa2     = spend2 / conv2 if conv2 > 0 and spend2 else None
            cvr2     = conv2 / clicks2 * 100 if clicks2 > 0 else None
            roas2    = val2 / spend2 if spend2 and spend2 > 0 and val2 > 0 else None
            row_vals = [
                safe(row.get("search_term")),
                round(float(spend2), 2) if spend2 else 0,
                clicks2,
                f"{ctr2:.2f}%",
                fmt_c(cpc2) if cpc2 else "—",
                f"{conv2:.1f}",
                f"{cvr2:.2f}%" if cvr2 else "—",
                fmt_c(cpa2) if cpa2 else "—",
                round(float(val2), 2) if val2 > 0 else None,
                f"{roas2:.1f}x" if roas2 else "—",
            ]
            for col, v in enumerate(row_vals, 1):
                c = wsp.cell(pr, col, v)
                c.fill = stripe
                c.font = Font(size=9)
                c.alignment = left(1) if col == 1 else center()
            pr += 1

        pr += 1  # gap between sections

    set_widths(wsp, {"A": 44, "B": 10, "C": 8, "D": 8, "E": 8, "F": 10,
                     "G": 12, "H": 10, "I": 10, "J": 12, "K": 8})
    wsp.freeze_panes = "A4"

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3 — All Terms (aggregated: one row per unique search term)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("All Terms")
    ws3.sheet_view.showGridLines = False

    agg_cols = [
        ("search_term",      "Search Term"),
        ("tier",             "Tier"),
        ("intent_category",  "Intent"),
        ("funnel_stage",     "Funnel Stage"),
        ("channels",         "Channels"),
        ("campaigns",        "Campaigns"),
        ("impressions",      "Impressions"),
        ("clicks",           "Clicks"),
        ("ctr",              "CTR"),
        ("cpc",              "CPC"),
        ("cost",             "Spend"),
        ("conversions",      "Conversions"),
        ("conv_value",       "Conv Value"),
        ("cpa",              "CPA"),
        ("roas",             "ROAS"),
        ("wasted_spend",     "Wasted Spend"),
    ]

    write_header_row(ws3, 1, [h for _, h in agg_cols], bg=NAVY, height=20)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(agg_cols))}1"

    for ri, (_, row) in enumerate(df_agg.iterrows(), 2):
        tier = row["tier"]
        stripe = hfill(LGRAY) if ri % 2 == 0 else hfill("FFFFFF")
        for ci, (key, _) in enumerate(agg_cols, 1):
            val = safe(row.get(key))
            if key in ("cost", "conv_value", "wasted_spend", "cpc") and val is not None:
                val = round(float(val), 2)
            elif key in ("cpa", "roas") and val is not None:
                val = round(float(val), 2)
            elif key == "ctr" and val is not None:
                val = round(float(val), 2)
            elif key in ("impressions", "clicks") and val is not None:
                val = int(float(val))
            elif key == "conversions" and val is not None:
                val = round(float(val), 1)
            c = ws3.cell(ri, ci, val)
            c.font = Font(size=9)
            c.alignment = left(1) if ci == 1 else center()
            if key == "tier":
                c.fill = tfill(tier)
                c.font = tfont(tier, size=9)
            else:
                c.fill = stripe

    set_widths(ws3, {"A": 44, "B": 16, "C": 14, "D": 14, "E": 26, "F": 14,
                     "G": 12, "H": 8,  "I": 8,  "J": 10, "K": 10, "L": 12, "M": 12, "N": 10, "O": 8, "P": 14})

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 4 — searchTerm_data (raw: one row per search term per campaign)
    # ══════════════════════════════════════════════════════════════════════════
    ws_raw = wb.create_sheet("searchTerm_data")
    ws_raw.sheet_view.showGridLines = False

    # Add cpc to raw df for searchTerm_data sheet
    df["cpc"] = df.apply(lambda r: r["cost"] / r["clicks"] if r["clicks"] > 0 else None, axis=1)

    a_cols = [
        ("search_term",     "Search Term"),
        ("tier",            "Tier"),
        ("intent_category", "Intent"),
        ("funnel_stage",    "Funnel Stage"),
        ("channel_type",    "Channel"),
        ("campaign",        "Campaign"),
        ("ad_group",        "Ad Group"),
        ("impressions",     "Impressions"),
        ("clicks",          "Clicks"),
        ("ctr",             "CTR"),
        ("cpc",             "CPC"),
        ("cost",            "Spend"),
        ("conversions",     "Conversions"),
        ("conv_value",      "Conv Value"),
        ("cpa",             "CPA"),
        ("roas",            "ROAS"),
        ("wasted_spend",    "Wasted Spend"),
    ]

    write_header_row(ws_raw, 1, [h for _, h in a_cols], bg=NAVY, height=20)
    ws_raw.freeze_panes = "A2"
    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(a_cols))}1"

    df_sorted = df.sort_values("cost", ascending=False).reset_index(drop=True)
    for ri, (_, row) in enumerate(df_sorted.iterrows(), 2):
        tier = row["tier"]
        stripe = hfill(LGRAY) if ri % 2 == 0 else hfill("FFFFFF")
        for ci, (key, _) in enumerate(a_cols, 1):
            val = safe(row.get(key))
            if key in ("cost", "conv_value", "wasted_spend", "cpc") and val is not None:
                val = round(float(val), 2)
            elif key in ("cpa", "roas", "ctr") and val is not None:
                val = round(float(val), 2)
            elif key in ("impressions", "clicks") and val is not None:
                val = int(float(val))
            elif key == "conversions" and val is not None:
                val = round(float(val), 1)
            c = ws_raw.cell(ri, ci, val)
            c.font = Font(size=9)
            c.alignment = left(1) if ci == 1 else center()
            if key == "tier":
                c.fill = tfill(tier)
                c.font = tfont(tier, size=9)
            else:
                c.fill = stripe

    set_widths(ws_raw, {"A": 44, "B": 16, "C": 14, "D": 14, "E": 16, "F": 32,
                        "G": 30, "H": 12, "I": 8,  "J": 8,  "K": 10, "L": 10, "M": 12, "N": 12, "O": 10, "P": 8, "Q": 14})

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 5 — Intent Breakdown
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Intent Breakdown")
    ws4.sheet_view.showGridLines = False

    intent_order_xl = ["Brand", "Competitors", "Product", "Generic"]

    _INTENT_BG = {
        "Brand":       "D9E1F2",
        "Competitors": "FCE4D6",
        "Product":     "E2EFDA",
        "Generic":     "EDEDED",
    }
    _INTENT_FG = {
        "Brand":       "17375E",
        "Competitors": "843C0C",
        "Product":     "375623",
        "Generic":     "444444",
    }
    _INTENT_LABEL = {
        "Brand":       "BRAND TERMS",
        "Competitors": "COMPETITOR TERMS",
        "Product":     "PRODUCT TERMS",
        "Generic":     "GENERIC TERMS",
    }

    ri4 = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws4.merge_cells(f"A{ri4}:K{ri4}")
    c = ws4.cell(ri4, 1, "INTENT BREAKDOWN")
    c.fill = hfill(NAVY); c.font = Font(bold=True, size=13, color=WHITE)
    c.alignment = center()
    ws4.row_dimensions[ri4].height = 28
    ri4 += 1

    ws4.merge_cells(f"A{ri4}:K{ri4}")
    c = ws4.cell(ri4, 1, f"{account_name}  |  {start_date} – {end_date}")
    c.fill = hfill(NAVY); c.font = Font(size=9, color="AAAAAA")
    c.alignment = center()
    ws4.row_dimensions[ri4].height = 16
    ri4 += 1
    ws4.row_dimensions[ri4].height = 8
    ri4 += 1

    # ── Summary table ─────────────────────────────────────────────────────────
    intent_sum_hdr = ["Intent Category", "Terms", "Spend", "Clicks", "CTR", "CPC",
                      "Conversions", "Conv Rate", "CPA", "Conv Value", "ROAS"]
    write_header_row(ws4, ri4, intent_sum_hdr, bg=COBALT, height=18)
    ri4 += 1

    for ck in intent_order_xl:
        sub = df_agg[df_agg["cluster_key"].str.startswith(ck)]
        if sub.empty:
            continue
        spend  = sub["cost"].sum()
        clicks = int(sub["clicks"].sum())
        impr   = sub["impressions"].sum()
        conv   = sub["conversions"].sum()
        val    = sub["conv_value"].sum()
        ctr    = clicks / impr * 100 if impr > 0 else 0
        cpc    = spend / clicks if clicks > 0 else None
        cpa    = spend / conv if conv > 0 else None
        cvr    = conv / clicks * 100 if clicks > 0 else None
        roas   = val / spend if spend > 0 and val > 0 else None

        bg = _INTENT_BG.get(ck, "FFFFFF")
        fg = _INTENT_FG.get(ck, "000000")
        vals = [
            ck, len(sub), fmt_c(spend), clicks,
            f"{ctr:.2f}%",
            fmt_c(cpc) if cpc else "—",
            f"{conv:.0f}",
            f"{cvr:.2f}%" if cvr else "—",
            fmt_c(cpa) if cpa else "—",
            fmt_c(val) if val > 0 else "—",
            f"{roas:.1f}x" if roas else "—",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws4.cell(ri4, ci, v)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(color=fg, size=9)
            c.alignment = left(1) if ci == 1 else center()
        ws4.row_dimensions[ri4].height = 18
        ri4 += 1

    ri4 += 1  # spacer

    # ── Per-intent detail sections ─────────────────────────────────────────────
    intent_detail_hdr = ["Search Term", "Tier", "Spend", "Clicks", "CTR", "CPC",
                         "Conversions", "Conv Rate", "CPA", "Conv Value", "ROAS"]

    for ck in intent_order_xl:
        sub = df_agg[df_agg["cluster_key"].str.startswith(ck)].sort_values("cost", ascending=False)
        if sub.empty:
            continue
        bg    = _INTENT_BG.get(ck, "FFFFFF")
        fg    = _INTENT_FG.get(ck, "000000")
        label = _INTENT_LABEL.get(ck, ck.upper())

        ws4.merge_cells(start_row=ri4, start_column=1, end_row=ri4, end_column=len(intent_detail_hdr))
        c = ws4.cell(ri4, 1, f"{label}  ({len(sub)} terms)")
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(bold=True, size=10, color=fg)
        c.alignment = left(1)
        ws4.row_dimensions[ri4].height = 20
        ri4 += 1

        write_header_row(ws4, ri4, intent_detail_hdr, bg=DGRAY, height=16)
        ri4 += 1

        for ri_idx, (_, row) in enumerate(sub.iterrows()):
            stripe  = hfill(LGRAY) if ri_idx % 2 == 0 else hfill("FFFFFF")
            spend2  = safe(row.get("cost"))
            clicks2 = int(float(safe(row.get("clicks")) or 0))
            impr2   = float(safe(row.get("impressions")) or 0)
            conv2   = float(safe(row.get("conversions")) or 0)
            val2    = float(safe(row.get("conv_value")) or 0)
            ctr2    = clicks2 / impr2 * 100 if impr2 > 0 else 0
            cpc2    = spend2 / clicks2 if clicks2 > 0 and spend2 else None
            cpa2    = spend2 / conv2 if conv2 > 0 and spend2 else None
            cvr2    = conv2 / clicks2 * 100 if clicks2 > 0 else None
            roas2   = val2 / spend2 if spend2 and spend2 > 0 and val2 > 0 else None
            tier2   = safe(row.get("tier")) or ""
            row_vals = [
                safe(row.get("search_term")),
                tier_labels.get(tier2, tier2),
                round(float(spend2), 2) if spend2 else 0,
                clicks2,
                f"{ctr2:.2f}%",
                fmt_c(cpc2) if cpc2 else "—",
                f"{conv2:.1f}",
                f"{cvr2:.2f}%" if cvr2 else "—",
                fmt_c(cpa2) if cpa2 else "—",
                round(float(val2), 2) if val2 > 0 else None,
                f"{roas2:.1f}x" if roas2 else "—",
            ]
            for ci, v in enumerate(row_vals, 1):
                c = ws4.cell(ri4, ci, v)
                c.fill = stripe
                c.font = Font(size=9)
                c.alignment = left(1) if ci <= 2 else center()
            ri4 += 1

        ri4 += 1  # gap between sections

    set_widths(ws4, {"A": 38, "B": 14, "C": 10, "D": 8, "E": 8, "F": 10,
                     "G": 12, "H": 10, "I": 10, "J": 12, "K": 8})
    ws4.freeze_panes = "A4"

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 6 — Funnel Breakdown
    # ══════════════════════════════════════════════════════════════════════════
    ws_funnel = wb.create_sheet("Funnel Breakdown")
    ws_funnel.sheet_view.showGridLines = False

    _FUNNEL_BG = {
        "Top Funnel":    "EDEDED",
        "Mid Funnel":    "FFEB9C",
        "Bottom Funnel": "C6EFCE",
    }
    _FUNNEL_FG = {
        "Top Funnel":    "595959",
        "Mid Funnel":    "7D6608",
        "Bottom Funnel": "375623",
    }
    funnel_stages = ["Bottom Funnel", "Mid Funnel", "Top Funnel"]

    rif = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws_funnel.merge_cells(f"A{rif}:K{rif}")
    c = ws_funnel.cell(rif, 1, "FUNNEL BREAKDOWN")
    c.fill = hfill(NAVY); c.font = Font(bold=True, size=13, color=WHITE)
    c.alignment = center()
    ws_funnel.row_dimensions[rif].height = 28
    rif += 1

    ws_funnel.merge_cells(f"A{rif}:K{rif}")
    c = ws_funnel.cell(rif, 1, f"{account_name}  |  {start_date} – {end_date}")
    c.fill = hfill(NAVY); c.font = Font(size=9, color="AAAAAA")
    c.alignment = center()
    ws_funnel.row_dimensions[rif].height = 16
    rif += 1
    ws_funnel.row_dimensions[rif].height = 8
    rif += 1

    # ── Summary table ─────────────────────────────────────────────────────────
    funnel_sum_hdr = ["Funnel Stage", "Terms", "Spend", "Clicks", "CTR", "CPC",
                      "Conversions", "Conv Rate", "CPA", "Conv Value", "ROAS"]
    write_header_row(ws_funnel, rif, funnel_sum_hdr, bg=COBALT, height=18)
    rif += 1

    for stage in funnel_stages:
        sub = df_agg[df_agg["funnel_stage"] == stage]
        if sub.empty:
            continue
        spend  = sub["cost"].sum()
        clicks = int(sub["clicks"].sum())
        impr   = sub["impressions"].sum()
        conv   = sub["conversions"].sum()
        val    = sub["conv_value"].sum()
        ctr    = clicks / impr * 100 if impr > 0 else 0
        cpc    = spend / clicks if clicks > 0 else None
        cpa    = spend / conv if conv > 0 else None
        cvr    = conv / clicks * 100 if clicks > 0 else None
        roas   = val / spend if spend > 0 and val > 0 else None
        bg = _FUNNEL_BG.get(stage, "FFFFFF")
        fg = _FUNNEL_FG.get(stage, DGRAY)
        vals = [
            stage, len(sub), fmt_c(spend), clicks,
            f"{ctr:.2f}%",
            fmt_c(cpc) if cpc else "—",
            f"{conv:.0f}",
            f"{cvr:.2f}%" if cvr else "—",
            fmt_c(cpa) if cpa else "—",
            fmt_c(val) if val > 0 else "—",
            f"{roas:.1f}x" if roas else "—",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws_funnel.cell(rif, ci, v)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(color=fg, size=9)
            c.alignment = left(1) if ci == 1 else center()
        ws_funnel.row_dimensions[rif].height = 18
        rif += 1

    rif += 1  # spacer

    # ── Per-stage detail sections ──────────────────────────────────────────────
    funnel_detail_hdr = ["Search Term", "Intent", "Spend", "Clicks", "CTR", "CPC",
                         "Conversions", "Conv Rate", "CPA", "Conv Value", "ROAS"]

    for stage in funnel_stages:
        sub = df_agg[df_agg["funnel_stage"] == stage].sort_values("cost", ascending=False)
        if sub.empty:
            continue
        bg = _FUNNEL_BG.get(stage, "FFFFFF")
        fg = _FUNNEL_FG.get(stage, DGRAY)

        ws_funnel.merge_cells(start_row=rif, start_column=1, end_row=rif, end_column=len(funnel_detail_hdr))
        c = ws_funnel.cell(rif, 1, f"{stage.upper()}  ({len(sub)} terms)")
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(bold=True, size=10, color=fg)
        c.alignment = left(1)
        ws_funnel.row_dimensions[rif].height = 20
        rif += 1

        write_header_row(ws_funnel, rif, funnel_detail_hdr, bg=DGRAY, height=16)
        rif += 1

        for ri_idx, (_, row) in enumerate(sub.iterrows()):
            stripe  = hfill(LGRAY) if ri_idx % 2 == 0 else hfill("FFFFFF")
            spend2  = safe(row.get("cost"))
            clicks2 = int(float(safe(row.get("clicks")) or 0))
            impr2   = float(safe(row.get("impressions")) or 0)
            conv2   = float(safe(row.get("conversions")) or 0)
            val2    = float(safe(row.get("conv_value")) or 0)
            ctr2    = clicks2 / impr2 * 100 if impr2 > 0 else 0
            cpc2    = spend2 / clicks2 if clicks2 > 0 and spend2 else None
            cpa2    = spend2 / conv2 if conv2 > 0 and spend2 else None
            cvr2    = conv2 / clicks2 * 100 if clicks2 > 0 else None
            roas2   = val2 / spend2 if spend2 and spend2 > 0 and val2 > 0 else None
            intent2 = safe(row.get("intent_category")) or ""
            row_vals = [
                safe(row.get("search_term")),
                intent2,
                round(float(spend2), 2) if spend2 else 0,
                clicks2,
                f"{ctr2:.2f}%",
                fmt_c(cpc2) if cpc2 else "—",
                f"{conv2:.1f}",
                f"{cvr2:.2f}%" if cvr2 else "—",
                fmt_c(cpa2) if cpa2 else "—",
                round(float(val2), 2) if val2 > 0 else None,
                f"{roas2:.1f}x" if roas2 else "—",
            ]
            for ci, v in enumerate(row_vals, 1):
                c = ws_funnel.cell(rif, ci, v)
                c.fill = stripe
                c.font = Font(size=9)
                c.alignment = left(1) if ci <= 2 else center()
            rif += 1

        rif += 1  # gap between sections

    set_widths(ws_funnel, {"A": 38, "B": 16, "C": 10, "D": 8, "E": 8, "F": 10,
                           "G": 12, "H": 10, "I": 10, "J": 12, "K": 8})
    ws_funnel.freeze_panes = "A4"

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet — Recommendations
    # ══════════════════════════════════════════════════════════════════════════
    ws_rec = wb.create_sheet("Recommendations")
    ws_rec.sheet_view.showGridLines = False

    _PRI_FILL = {"High": "FFC7CE", "Medium": "FFEB9C", "Low": "C6EFCE"}
    _PRI_FONT = {"High": "9C0006", "Medium": "7D6608", "Low": "375623"}

    rr = 1
    ws_rec.merge_cells(f"A{rr}:D{rr}")
    c = ws_rec.cell(rr, 1, "RECOMMENDATIONS")
    c.fill = hfill(NAVY); c.font = Font(bold=True, size=13, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_rec.row_dimensions[rr].height = 28
    rr += 1

    ws_rec.merge_cells(f"A{rr}:D{rr}")
    c = ws_rec.cell(rr, 1, f"{account_name}  |  {start_date} – {end_date}")
    c.fill = hfill(NAVY); c.font = Font(size=9, color="AAAAAA")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_rec.row_dimensions[rr].height = 16
    rr += 1
    ws_rec.row_dimensions[rr].height = 6
    rr += 1

    write_header_row(ws_rec, rr, ["Priority", "Category", "Finding", "Recommendation"], bg=COBALT, height=18)
    rr += 1

    if not recs:
        ws_rec.merge_cells(f"A{rr}:D{rr}")
        c = ws_rec.cell(rr, 1, "No significant issues detected.")
        c.font = Font(size=9, color=DGRAY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_rec.row_dimensions[rr].height = 18
    else:
        for idx, rec in enumerate(recs):
            stripe = hfill(LGRAY) if idx % 2 == 0 else hfill("FFFFFF")
            pri = rec["priority"]
            for ci, val in enumerate([pri, rec["category"], rec["finding"], rec["action"]], 1):
                c = ws_rec.cell(rr, ci, val)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
                c.font = Font(size=9)
                c.fill = stripe
            # Colour the Priority cell
            pc = ws_rec.cell(rr, 1)
            pc.fill = PatternFill("solid", fgColor=_PRI_FILL.get(pri, "FFFFFF"))
            pc.font = Font(bold=True, size=9, color=_PRI_FONT.get(pri, "000000"))
            pc.alignment = Alignment(horizontal="center", vertical="top")
            ws_rec.row_dimensions[rr].height = max(30, min(80, max(len(rec["finding"]), len(rec["action"])) // 3))
            rr += 1

    set_widths(ws_rec, {"A": 11, "B": 22, "C": 52, "D": 62})

    def sprawl_rows(df_group):
        """Build list of dicts for a groupby result."""
        rows_out = []
        for name, grp in df_group:
            n         = len(grp)
            spend     = grp["cost"].sum()
            prof_s    = grp.loc[grp["tier"] == "profitable",   "cost"].sum()
            unprof_s  = grp.loc[grp["tier"] == "unprofitable", "cost"].sum()
            unt_s     = grp.loc[grp["tier"] == "untested",     "cost"].sum()
            waste_s   = grp.loc[grp["tier"] == "waste",        "cost"].sum()
            sprawl_pct = (unt_s + waste_s) / spend * 100 if spend > 0 else 0
            # Intent percentages (% of terms, not spend)
            ck = grp.get("cluster_key", pd.Series(dtype=str))
            pct_brand   = (ck == "Brand").sum()   / n * 100 if n else 0
            pct_generic = (ck == "Generic").sum() / n * 100 if n else 0
            rows_out.append({
                "name":         name,
                "terms":        n,
                "spend":        spend,
                "profitable":   (grp["tier"] == "profitable").sum(),
                "prof_spend":   prof_s,
                "unprofitable": (grp["tier"] == "unprofitable").sum(),
                "unprof_spend": unprof_s,
                "untested":     (grp["tier"] == "untested").sum(),
                "unt_spend":    unt_s,
                "waste":        (grp["tier"] == "waste").sum(),
                "waste_spend":  waste_s,
                "sprawl_pct":   sprawl_pct,
                "pct_brand":    pct_brand,
                "pct_generic":  pct_generic,
            })
        return sorted(rows_out, key=lambda x: x["spend"], reverse=True)

    def write_sprawl_section(ws, start_row, title, rows_data):
        """Write a titled section of sprawl rows. Returns next available row."""
        r = start_row
        # Section title
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(SPRAWL_COLS))  # noqa: dynamic
        tc = ws.cell(r, 1, title)
        tc.fill  = hfill(NAVY)
        tc.font  = Font(color="FFFFFF", bold=True, size=10)
        tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18
        r += 1
        # Header
        write_header_row(ws, r, SPRAWL_COLS, bg="2B303B", height=16)
        r += 1
        # Data rows
        for rd in rows_data:
            stripe = hfill(LGRAY) if r % 2 == 0 else hfill("FFFFFF")
            for ci, key in enumerate(SPRAWL_COL_KEYS, 1):
                val = rd[key]
                if key in ("spend", "prof_spend", "unprof_spend", "unt_spend", "waste_spend"):
                    val = round(float(val), 2)
                elif key in ("sprawl_pct", "pct_brand", "pct_generic"):
                    val = round(float(val), 1)
                c = ws.cell(r, ci, val)
                c.font      = Font(size=9)
                c.alignment = left(1) if ci == 1 else center()
                c.fill      = stripe
                # Sprawl %: red=bad, yellow=medium, green=good
                if key == "sprawl_pct" and isinstance(val, (int, float)):
                    if val >= 80:
                        c.fill = hfill("FFC7CE"); c.font = Font(color="9C0006", size=9)
                    elif val >= 50:
                        c.fill = hfill("FFEB9C"); c.font = Font(color="7D6608", size=9)
                    else:
                        c.fill = hfill("C6EFCE"); c.font = Font(color="375623", size=9)
                # Generic %: high = opportunity for negatives (orange tint above 30%)
                elif key == "pct_generic" and isinstance(val, (int, float)):
                    if val >= 30:
                        c.fill = hfill("FCE4D6"); c.font = Font(color="843C0C", size=9)
                    elif val >= 15:
                        c.fill = hfill("FFEB9C"); c.font = Font(color="7D6608", size=9)
            r += 1
        return r + 1  # spacer

    # Sprawl Analysis tab removed — high sprawl is expected for eCommerce
    # accounts running Shopping campaigns and is not actionable.

    # ══════════════════════════════════════════════════════════════════════════
    # README tab
    # ══════════════════════════════════════════════════════════════════════════
    ws_readme = wb.create_sheet("README")
    ws_readme.sheet_view.showGridLines = False
    ws_readme.column_dimensions["A"].width = 22
    ws_readme.column_dimensions["B"].width = 80

    def readme_title(ws, r, text):
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(r, 1, text)
        c.fill = hfill(NAVY); c.font = Font(bold=True, size=13, color=WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 28
        return r + 1

    def readme_section(ws, r, text):
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(r, 1, text)
        c.fill = hfill("E8F2FF"); c.font = Font(bold=True, size=10, color=DGRAY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 20
        return r + 1

    def readme_row(ws, r, label, value, bold_label=False):
        cl = ws.cell(r, 1, label)
        cl.font = Font(bold=bold_label, size=9, color=DGRAY)
        cl.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
        cv = ws.cell(r, 2, value)
        cv.font = Font(size=9)
        cv.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(15, min(60, len(str(value)) // 4))
        return r + 1

    def readme_spacer(ws, r):
        ws.row_dimensions[r].height = 6
        return r + 1

    rr = 1
    rr = readme_title(ws_readme, rr, "SEARCH TERM AUDIT — HOW TO READ THIS FILE")
    rr = readme_spacer(ws_readme, rr)

    rr = readme_section(ws_readme, rr, "WHAT THIS FILE IS")
    rr = readme_row(ws_readme, rr, "", "This report classifies every search term that triggered your ads into profitability tiers. It tells you which terms are earning their budget, which are wasting it, and which haven't been tested yet.")
    rr = readme_spacer(ws_readme, rr)

    rr = readme_section(ws_readme, rr, "HOW TERMS ARE CLASSIFIED")
    rr = readme_row(ws_readme, rr, "Profitable", "Spent ≥ threshold AND converting at or below your target CPA (or above target ROAS). Keep these.")
    rr = readme_row(ws_readme, rr, "Unprofitable", "Spent ≥ threshold with zero conversions, or CPA > 3× target. These are candidates for negatives.")
    rr = readme_row(ws_readme, rr, "Untested", "Spent less than the threshold. Not enough data to judge. Lower the threshold with --min-spend if you want more terms evaluated.")
    rr = readme_row(ws_readme, rr, "Waste", "Structurally irrelevant regardless of spend — e.g. job searches, Reddit, DIY queries. Negatives recommended.")
    rr = readme_spacer(ws_readme, rr)

    rr = readme_section(ws_readme, rr, "SPEND THRESHOLD")
    rr = readme_row(ws_readme, rr, "", f"Set to 2× target CPA by default. A term must reach this spend level before being called profitable or unprofitable. This avoids false positives from low-data terms. Use --min-spend to override.")
    rr = readme_spacer(ws_readme, rr)

    rr = readme_section(ws_readme, rr, "INTENT CATEGORIES")
    rr = readme_row(ws_readme, rr, "Brand", "Searches that include your brand name or product names.")
    rr = readme_row(ws_readme, rr, "Competitors", "Searches for competing brands or products.")
    rr = readme_row(ws_readme, rr, "Product", "Category or product-type searches without a specific brand.")
    rr = readme_row(ws_readme, rr, "Generic", "Broad informational or non-specific queries.")
    rr = readme_spacer(ws_readme, rr)

    rr = readme_section(ws_readme, rr, "FUNNEL STAGES")
    rr = readme_row(ws_readme, rr, "Bottom Funnel", "High purchase intent — e.g. 'buy', 'order', brand + product.")
    rr = readme_row(ws_readme, rr, "Mid Funnel", "Comparison or consideration stage — e.g. 'best', 'vs', 'review'.")
    rr = readme_row(ws_readme, rr, "Top Funnel", "Awareness or research stage — e.g. 'what is', 'how to', broad category.")
    rr = readme_spacer(ws_readme, rr)

    rr = readme_section(ws_readme, rr, "TABS IN THIS FILE")
    rr = readme_row(ws_readme, rr, "Summary", "Top-level overview: total spend, waste, ROAS, and key breakdowns.")
    rr = readme_row(ws_readme, rr, "Profitability Summary", "Full list of profitable and unprofitable terms with spend and CPA.")
    rr = readme_row(ws_readme, rr, "Recommendations", "Data-driven recommendations ranked by priority. Each finding includes the observation and a specific action.")
    rr = readme_row(ws_readme, rr, "Intent Breakdown", "Performance by intent category: Brand, Competitors, Product, Generic.")
    rr = readme_row(ws_readme, rr, "Funnel Breakdown", "Performance by funnel stage: Bottom, Mid, Top.")
    rr = readme_row(ws_readme, rr, "All Terms", "Every search term with its tier, intent, funnel stage, and metrics.")
    rr = readme_row(ws_readme, rr, "searchTerm_data", "Raw data pulled from the Google Ads API.")

    # ── Reorder sheets ────────────────────────────────────────────────────────
    desired_order = ["README", "Recommendations", "Summary", "Profitability Summary",
                     "Intent Breakdown", "Funnel Breakdown", "All Terms", "searchTerm_data"]
    sheet_names = [ws.title for ws in wb.worksheets]
    for i, name in enumerate(desired_order):
        if name in sheet_names:
            wb.move_sheet(name, offset=wb.worksheets.index(wb[name]) * -1 + i)

    # ── Save ──────────────────────────────────────────────────────────────────
    xlsx_path = output_dir / f"{today}-{account_key}-audit.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="B6 Search Term Profitability Audit",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python b6_search_term_audit.py --account=bsk --days=30 --target-cpa=50
  python b6_search_term_audit.py --account=bsk --days=90 --target-cpa=auto
  python b6_search_term_audit.py --account=bsk --start=2026-01-01 --end=2026-01-31 --target-cpa=50
        """
    )
    parser.add_argument("--account", required=True, help="Account alias from accounts.json")
    parser.add_argument("--days", type=int, default=30, help="Date range in days (default: 30)")
    parser.add_argument("--start", help="Custom start date YYYY-MM-DD")
    parser.add_argument("--end", help="Custom end date YYYY-MM-DD")
    parser.add_argument("--target-cpa", default=None, help="Target CPA in account currency, or 'auto'")
    parser.add_argument("--target-roas", type=float, default=None, help="Target ROAS as percentage, e.g. 160 for 160%%")
    parser.add_argument("--min-spend", type=float, help="Override minimum spend threshold")
    parser.add_argument("--currency", default=None, help="Currency symbol (default: from accounts.json)")
    parser.add_argument("--min-impressions", type=int, default=10, help="Minimum impressions to include a search term (default: 10)")
    parser.add_argument("--no-clusters", action="store_true", help="Skip intent classification")
    parser.add_argument("--use-cache", action="store_true", help="Load raw search term data from previous fetch (skip API call)")
    parser.add_argument("--output", default=None, help="Output directory path")
    return parser.parse_args()


def main():
    args = parse_args()
    config = load_config()
    accounts = load_accounts()

    # ── Resolve account ───────────────────────────────────────────────────────
    try:
        account_key, account = resolve_account(args.account, accounts)
    except ValueError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    customer_id = account["id"]
    login_customer_id = account.get("login_customer_id", customer_id)
    currency = args.currency or account.get("currency_symbol", "$")
    account_name = account.get("name", account_key)

    # ── Date range ────────────────────────────────────────────────────────────
    start_date, end_date = get_date_range(
        days=args.days,
        start=args.start,
        end=args.end,
        timezone=account.get("timezone", "UTC")
    )

    # ── Fetch data ────────────────────────────────────────────────────────────
    raw_cache_path = (Path(args.output) if args.output else DATA_DIR / account_key) / \
                     f"raw-cache-{start_date}-{end_date}.csv"

    if args.use_cache and raw_cache_path.exists():
        print(f"\nLoading cached search terms for {account_name} ({start_date} – {end_date})...")
        df = pd.read_csv(raw_cache_path)
    else:
        print(f"\nFetching search terms for {account_name} ({start_date} – {end_date})...")
        df = fetch_search_terms(customer_id, login_customer_id, start_date, end_date,
                                min_impressions=args.min_impressions)
        raw_cache_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(raw_cache_path, index=False)

    if df.empty:
        print("No search terms found for this period.")
        sys.exit(0)

    print(f"  {len(df):,} search terms fetched.")

    # ── Target metric ─────────────────────────────────────────────────────────
    target_cpa = None
    target_roas = None

    if args.target_roas:
        # Convert percentage to decimal (160% → 1.60)
        target_roas = args.target_roas / 100.0
        # Spend threshold: avg order value / target_roas * multiplier
        avg_order_value = calculate_avg_order_value(df)
        if avg_order_value:
            spend_threshold = (avg_order_value / target_roas) * config["thresholds"]["spend_multiplier"]
            print(f"  Target ROAS: {args.target_roas:.0f}%  |  Avg order value: {currency}{avg_order_value:.2f}")
        else:
            spend_threshold = 50.0  # sensible fallback
            print(f"  Target ROAS: {args.target_roas:.0f}%  |  No conv value data — using ${spend_threshold} spend threshold")
    elif args.target_cpa:
        if args.target_cpa == "auto":
            target_cpa = calculate_account_cpa(df)
            print(f"  Target CPA calculated from data: {currency}{target_cpa:.2f}")
        else:
            try:
                target_cpa = float(args.target_cpa)
            except ValueError:
                print("ERROR: --target-cpa must be a number or 'auto'")
                sys.exit(1)
        spend_threshold = target_cpa * config["thresholds"]["spend_multiplier"]
    else:
        print("ERROR: provide --target-cpa or --target-roas")
        sys.exit(1)

    # Override spend threshold if --min-spend provided
    if args.min_spend:
        spend_threshold = args.min_spend

    print(f"  Spend threshold: {currency}{spend_threshold:.2f}")

    # ── Build intent categories ───────────────────────────────────────────────
    if args.no_clusters:
        df["cluster_key"]     = df["search_term"]
        df["cluster_size"]    = 1
        df["intent_category"] = None
        df["funnel_stage"]    = None
    else:
        df = build_intent_clusters(df, account_key, account, config)

    df["cluster"] = df["cluster_key"]

    # ── Classify ──────────────────────────────────────────────────────────────
    df = classify_terms(
        df,
        spend_threshold=spend_threshold,
        target_cpa=target_cpa,
        target_roas=target_roas,
        high_cpa_multiplier=config["thresholds"]["high_cpa_multiplier"],
        waste_patterns=config["structural_waste"],
        strip_modifiers=config["strip_modifiers"],
    )

    # ── Report ────────────────────────────────────────────────────────────────
    print_report(df, account_name, start_date, end_date, target_cpa, currency, config,
                 spend_threshold=spend_threshold, target_roas=target_roas)

    # ── Export ────────────────────────────────────────────────────────────────
    output_dir = args.output or DATA_DIR / account_key
    xlsx_path = export_excel(
        df, account_name, account_key, start_date, end_date,
        target_cpa, target_roas, spend_threshold, currency, config, output_dir
    )

    print(f"  Report → {xlsx_path}")
    print("═" * 68)
    print()


if __name__ == "__main__":
    main()
