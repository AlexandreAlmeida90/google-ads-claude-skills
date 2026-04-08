#!/usr/bin/env python3
"""
Generate Google Ads RSA Copy Excel file — vertical layout, clean design.

Usage:
    python generate-excel.py --input copy.json --output "RSA Copy - Client.xlsx"

Input JSON format:
{
  "client": "Vital Planet",
  "campaign": "Search - NonBrand",
  "ad_group": "Women's Probiotics",
  "final_url": "https://...",
  "headlines": ["H1", "H2", ... (15 total)],
  "headline_types": ["Key Features", "Key Features", ...],
  "descriptions": ["D1", "D2", "D3", "D4"],
  "description_types": ["Key Features / Benefits / USP", ...],
  "path1": "probiotics",
  "path2": "women"
}
"""

import json
import sys
import argparse
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.styles.differential import DifferentialStyle
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY        = 'FF1A2744'   # Section header bg
WHITE       = 'FFFFFFFF'
YELLOW_BG   = 'FFFFF2CC'   # Column sub-header bg
HL_STRIPE   = 'FFF0F4FF'   # Headline row stripe (light blue)
DESC_STRIPE = 'FFFFF8F0'   # Description row stripe (light warm)
GREEN_OK    = 'FF34A853'   # Status OK text
RED_OVER    = 'FFEA4335'   # Status OVER text
RED_FILL    = 'FFFCE8E6'   # Status OVER cell bg
GRAY_BG     = 'FFF8F9FA'   # Info block bg
GRAY_BORDER = 'FFD0D5DD'   # Border color
MED_GRAY    = 'FF6B7280'   # Muted text
DARK_TEXT   = 'FF111827'   # Primary text

# ── Headline type defaults ────────────────────────────────────────────────────
DEFAULT_HL_TYPES = [
    'Key Features',
    'Key Features',
    'Key Features',
    'Benefits',
    'Benefits',
    'Benefits',
    'Brand',
    'Social Proof',
    'Social Proof',
    'Target Audience',
    'USP',
    'USP',
    'Offer / CTA',
    'Offer / CTA',
    'Seasonal / Evergreen',
]

DEFAULT_DESC_TYPES = [
    'Key Features / Benefits / USP',
    'Key Features / Benefits / USP',
    'Social Proof / Authority',
    'Offer / CTA / Seasonal',
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def fc(hex8):
    """PatternFill from 8-char hex."""
    c = hex8 if len(hex8) == 8 else 'FF' + hex8.lstrip('#')
    return PatternFill(start_color=c, end_color=c, fill_type='solid')

def border(color=GRAY_BORDER, style='thin'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color=GRAY_BORDER):
    s = Side(style='thin', color=color)
    return Border(bottom=s)

def font(bold=False, size=10, color=DARK_TEXT, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name='Calibri')

def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def write(ws, row, col, value, bg=None, ft=None, al=None, bd=None):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:  cell.fill      = fc(bg)
    if ft:  cell.font      = ft
    if al:  cell.alignment = al
    if bd:  cell.border    = bd
    return cell

def merge_write(ws, row, col_start, col_end, value, bg=None, ft=None, al=None, bd=None):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    if bg:  cell.fill      = fc(bg)
    if ft:  cell.font      = ft
    if al:  cell.alignment = al
    if bd:  cell.border    = bd
    return cell

def apply_row_border(ws, row, col_start, col_end, color=GRAY_BORDER):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        existing = cell.border
        s = Side(style='thin', color=color)
        cell.border = Border(
            left=existing.left, right=existing.right,
            top=existing.top, bottom=s
        )


# ── Main builder ──────────────────────────────────────────────────────────────
def build(wb, data):
    ws = wb.active
    ws.title = "RSA Copy"

    # ── Column widths ─────────────────────────────────────────────────────────
    # A: spacer | B: # | C: Type | D: Copy text | E: Chars | F: Status
    ws.column_dimensions['A'].width = 1.5
    ws.column_dimensions['B'].width = 4.0
    ws.column_dimensions['C'].width = 28.0
    ws.column_dimensions['D'].width = 52.0
    ws.column_dimensions['E'].width = 7.0
    ws.column_dimensions['F'].width = 8.0
    ws.column_dimensions['G'].width = 1.5   # right spacer

    row = 1

    # ── ROW 1: top spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 10
    row += 1

    # ── ROW 2: Title bar ─────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 36
    merge_write(ws, row, 2, 6,
                f"Google Ads RSA Copy  |  {data.get('client', '')}",
                bg=NAVY,
                ft=font(bold=True, size=14, color=WHITE),
                al=align('left', 'center'))
    row += 1

    # ── ROW 3: Meta info block ────────────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    meta = (
        f"Campaign: {data.get('campaign', '—')}   |   "
        f"Ad Group: {data.get('ad_group', '—')}   |   "
        f"Generated: {datetime.today().strftime('%b %d, %Y')}"
    )
    merge_write(ws, row, 2, 6, meta,
                bg=GRAY_BG,
                ft=font(size=9, color=MED_GRAY),
                al=align('left', 'center'))
    row += 1

    # ── ROW 4: URL row ────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 16
    url = data.get('final_url', '')
    if url:
        merge_write(ws, row, 2, 6, f"Final URL: {url}",
                    bg=GRAY_BG,
                    ft=font(size=9, color='FF0070CC'),
                    al=align('left', 'center'))
    row += 1

    # ── ROW 5: spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 10
    row += 1

    # ── HEADLINES SECTION ─────────────────────────────────────────────────────

    # Section header
    ws.row_dimensions[row].height = 26
    merge_write(ws, row, 2, 6, "HEADLINES  (15 total  |  max 30 characters each)",
                bg=NAVY,
                ft=font(bold=True, size=11, color=WHITE),
                al=align('left', 'center'))
    row += 1

    # Column headers
    ws.row_dimensions[row].height = 18
    for col, (val, a) in enumerate([
        ('#',      'center'),
        ('Type',   'left'),
        ('Headline Copy',   'left'),
        ('Chars',  'center'),
        ('Status', 'center'),
    ], start=2):
        write(ws, row, col, val,
              bg=YELLOW_BG,
              ft=font(bold=True, size=9, color=DARK_TEXT),
              al=align(a, 'center'),
              bd=border())
    row += 1

    headlines     = data.get('headlines', [])
    hl_types      = data.get('headline_types', DEFAULT_HL_TYPES)

    hl_data_start = row
    for i in range(15):
        ws.row_dimensions[row].height = 22
        hl_text  = headlines[i] if i < len(headlines) else ''
        hl_type  = hl_types[i]  if i < len(hl_types)  else ''
        row_bg   = HL_STRIPE if i % 2 == 0 else WHITE

        # # col
        write(ws, row, 2, i + 1,
              bg=row_bg, ft=font(size=9, color=MED_GRAY), al=align('center'), bd=border(style='hair'))

        # Type col
        write(ws, row, 3, hl_type,
              bg=row_bg, ft=font(size=9, color=MED_GRAY, italic=True), al=align('left'), bd=border(style='hair'))

        # Copy text col
        hl_cell = write(ws, row, 4, hl_text,
              bg=row_bg, ft=font(size=10, color=DARK_TEXT), al=align('left'), bd=border(style='hair'))

        # Chars col — LEN formula
        char_cell = ws.cell(row=row, column=5)
        char_cell.value = f'=LEN(D{row})'
        char_cell.fill  = fc(row_bg)
        char_cell.font  = font(size=9, color=DARK_TEXT)
        char_cell.alignment = align('center')
        char_cell.border = border(style='hair')

        # Status col — formula-based
        status_cell = ws.cell(row=row, column=6)
        status_cell.value = f'=IF(LEN(D{row})>30,"OVER","OK")'
        status_cell.font  = font(size=9, bold=True)
        status_cell.alignment = align('center')
        status_cell.border = border(style='hair')

        row += 1

    hl_data_end = row - 1

    # Apply conditional formatting: OVER = red bg, OK = green text
    red_font  = Font(bold=True, color=RED_OVER,  name='Calibri', size=9)
    grn_font  = Font(bold=True, color=GREEN_OK,  name='Calibri', size=9)
    red_fill  = fc(RED_FILL)

    ws.conditional_formatting.add(
        f'F{hl_data_start}:F{hl_data_end}',
        CellIsRule(operator='equal', formula=['"OVER"'],
                   font=red_font, fill=red_fill)
    )
    ws.conditional_formatting.add(
        f'F{hl_data_start}:F{hl_data_end}',
        CellIsRule(operator='equal', formula=['"OK"'], font=grn_font)
    )

    # ── spacer ────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    row += 1

    # ── DESCRIPTIONS SECTION ──────────────────────────────────────────────────

    # Section header
    ws.row_dimensions[row].height = 26
    merge_write(ws, row, 2, 6, "DESCRIPTIONS  (4 total  |  max 90 characters each)",
                bg=NAVY,
                ft=font(bold=True, size=11, color=WHITE),
                al=align('left', 'center'))
    row += 1

    # Column headers
    ws.row_dimensions[row].height = 18
    for col, (val, a) in enumerate([
        ('#',     'center'),
        ('Type',  'left'),
        ('Description Copy', 'left'),
        ('Chars', 'center'),
        ('Status','center'),
    ], start=2):
        write(ws, row, col, val,
              bg=YELLOW_BG,
              ft=font(bold=True, size=9, color=DARK_TEXT),
              al=align(a, 'center'),
              bd=border())
    row += 1

    descriptions  = data.get('descriptions', [])
    desc_types    = data.get('description_types', DEFAULT_DESC_TYPES)

    desc_data_start = row
    for i in range(4):
        ws.row_dimensions[row].height = 38
        desc_text = descriptions[i] if i < len(descriptions) else ''
        desc_type = desc_types[i]   if i < len(desc_types)   else ''
        row_bg    = DESC_STRIPE if i % 2 == 0 else WHITE

        write(ws, row, 2, i + 1,
              bg=row_bg, ft=font(size=9, color=MED_GRAY), al=align('center', 'top'), bd=border(style='hair'))

        write(ws, row, 3, desc_type,
              bg=row_bg, ft=font(size=9, color=MED_GRAY, italic=True), al=align('left', 'top'), bd=border(style='hair'))

        write(ws, row, 4, desc_text,
              bg=row_bg, ft=font(size=10, color=DARK_TEXT), al=align('left', 'top', wrap=True), bd=border(style='hair'))

        char_cell = ws.cell(row=row, column=5)
        char_cell.value     = f'=LEN(D{row})'
        char_cell.fill      = fc(row_bg)
        char_cell.font      = font(size=9, color=DARK_TEXT)
        char_cell.alignment = align('center', 'top')
        char_cell.border    = border(style='hair')

        status_cell = ws.cell(row=row, column=6)
        status_cell.value     = f'=IF(LEN(D{row})>90,"OVER","OK")'
        status_cell.font      = font(size=9, bold=True)
        status_cell.alignment = align('center', 'top')
        status_cell.border    = border(style='hair')

        row += 1

    desc_data_end = row - 1

    ws.conditional_formatting.add(
        f'F{desc_data_start}:F{desc_data_end}',
        CellIsRule(operator='equal', formula=['"OVER"'],
                   font=red_font, fill=red_fill)
    )
    ws.conditional_formatting.add(
        f'F{desc_data_start}:F{desc_data_end}',
        CellIsRule(operator='equal', formula=['"OK"'], font=grn_font)
    )

    # ── spacer ────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    row += 1

    # ── AD PATHS SECTION ──────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 26
    merge_write(ws, row, 2, 6, "AD PATHS  (max 15 characters each)",
                bg=NAVY,
                ft=font(bold=True, size=11, color=WHITE),
                al=align('left', 'center'))
    row += 1

    for label, value in [('Path 1', data.get('path1', '')), ('Path 2', data.get('path2', ''))]:
        ws.row_dimensions[row].height = 22
        write(ws, row, 2, '',      bg=WHITE)
        write(ws, row, 3, label,   bg=WHITE, ft=font(size=10, bold=True, color=DARK_TEXT), al=align('left'), bd=border(style='hair'))
        write(ws, row, 4, value,   bg=WHITE, ft=font(size=10, color=DARK_TEXT), al=align('left'), bd=border(style='hair'))

        char_cell = ws.cell(row=row, column=5)
        char_cell.value     = f'=LEN(D{row})'
        char_cell.fill      = fc(WHITE)
        char_cell.font      = font(size=9)
        char_cell.alignment = align('center')
        char_cell.border    = border(style='hair')

        status_cell = ws.cell(row=row, column=6)
        status_cell.value     = f'=IF(LEN(D{row})>15,"OVER","OK")'
        status_cell.font      = font(size=9, bold=True)
        status_cell.alignment = align('center')
        status_cell.border    = border(style='hair')

        row += 1

    ws.conditional_formatting.add(
        f'F{row-2}:F{row-1}',
        CellIsRule(operator='equal', formula=['"OVER"'], font=red_font, fill=red_fill)
    )
    ws.conditional_formatting.add(
        f'F{row-2}:F{row-1}',
        CellIsRule(operator='equal', formula=['"OK"'], font=grn_font)
    )

    # ── bottom spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 16

    # ── Freeze header rows ────────────────────────────────────────────────────
    ws.freeze_panes = 'B7'

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = 'portrait'


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Generate RSA Copy Excel')
    parser.add_argument('--input',  required=True)
    parser.add_argument('--output', required=True)
    args = parser.parse_args()

    with open(args.input, 'r', encoding='utf-8') as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    build(wb, data)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"Excel saved: {out}")


if __name__ == '__main__':
    main()
